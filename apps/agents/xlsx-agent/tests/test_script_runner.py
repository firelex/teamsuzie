"""Tests for the script runner service."""

from pathlib import Path

import openpyxl
import pytest

from app.config import settings
from app.services.script_runner import run_script, write_script


def test_write_script_drops_library_next_to_script(tmp_path: Path):
    job_dir = tmp_path / "job1"
    code = "# empty\n"
    path = write_script(job_dir, code)
    assert path.exists()
    assert path.name == "script.py"
    assert (job_dir / "xlsx_builder.py").exists()


def test_write_script_handles_relative_job_dir(tmp_path: Path, monkeypatch):
    """Regression: relative job_dir should be absolutized."""
    monkeypatch.chdir(tmp_path)
    job_dir = Path("relative-job")
    path = write_script(job_dir, "# empty\n")
    assert path.is_absolute()
    assert path.exists()


def test_run_script_missing_script_returns_error(tmp_path: Path):
    job_dir = tmp_path / "job-no-script"
    job_dir.mkdir()
    result = run_script(job_dir)
    assert result.success is False
    assert "No script.py" in (result.error or "")


def test_run_script_happy_path_produces_xlsx(tmp_path: Path):
    job_dir = tmp_path / "job-happy"
    code = '''
import sys
from xlsx_builder import XLSXBuilder, Fmt

class Tiny(XLSXBuilder):
    def build(self):
        ws = self.add_sheet("Hello")
        self.title_row(ws, "Hi", "A", "B", row=1)
        self.s.label(ws["A3"], "Revenue")
        self.s.input(ws["B3"], value=1000, fmt=Fmt.CURRENCY)

if __name__ == "__main__":
    wb = Tiny()
    wb.build()
    wb.save(sys.argv[1])
'''
    write_script(job_dir, code)
    result = run_script(job_dir)
    assert result.success is True, result.error
    assert result.output_path is not None
    assert result.output_path.exists()
    assert result.output_path.name == "result.xlsx"

    wb = openpyxl.load_workbook(result.output_path)
    assert wb.sheetnames == ["Hello"]
    assert wb["Hello"]["B3"].value == 1000


def test_run_script_syntax_error_returns_failure(tmp_path: Path):
    job_dir = tmp_path / "job-bad"
    write_script(job_dir, "this is not valid python\n")
    result = run_script(job_dir)
    assert result.success is False
    assert "SyntaxError" in result.stderr or result.error


def test_run_script_no_output_file_returns_error(tmp_path: Path):
    """Script that completes but never writes sys.argv[1]."""
    job_dir = tmp_path / "job-no-output"
    write_script(job_dir, 'print("hello")\n')
    result = run_script(job_dir)
    assert result.success is False
    assert "did not write" in (result.error or "").lower()


def test_run_script_runtime_error_rollback(tmp_path: Path):
    """Script crashes mid-execution — no file, clear error."""
    job_dir = tmp_path / "job-crash"
    write_script(job_dir, "raise RuntimeError('boom')\n")
    result = run_script(job_dir)
    assert result.success is False
    assert "boom" in result.stderr
    assert result.output_path is None


def test_run_script_timeout(tmp_path: Path, monkeypatch):
    monkeypatch.setattr(settings, "script_timeout_seconds", 1)
    job_dir = tmp_path / "job-slow"
    write_script(job_dir, "import time; time.sleep(5)\n")
    result = run_script(job_dir)
    assert result.success is False
    assert "timeout" in (result.error or "").lower()
