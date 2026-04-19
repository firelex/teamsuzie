"""
xlsx_builder/templates/data_report.py
--------------------------------------
Data Table / Report template.

Produces a professional, ready-to-fill report workbook with three sheets:
  1. Cover       — title, metadata, legend
  2. Data        — column-header table with frozen panes, alternating row shading,
                   a totals row, and optional formula columns
  3. Summary     — auto-generated KPI tiles and a section for charts/notes

Usage:
    from data_report import DataReport
    from xlsx_builder import Fmt

    report = DataReport()
    report.build(
        title="Q2 Sales Report",
        subtitle="Regional Performance by Product Line",
        columns=[
            # (header,          width,  fmt,          total_func)
            ("Region",          20,     Fmt.TEXT,      None),
            ("Product Line",    22,     Fmt.TEXT,      None),
            ("Units Sold",      14,     Fmt.INTEGER,   "SUM"),
            ("Revenue ($)",     16,     Fmt.CURRENCY,  "SUM"),
            ("Avg Price ($)",   16,     Fmt.CURRENCY,  "AVERAGE"),
            ("Margin %",        12,     Fmt.PCT,       "AVERAGE"),
        ],
        data_rows=20,        # number of blank input rows to pre-populate
        metadata={
            "Prepared By":  "",
            "Department":   "",
            "Period":       "",
            "Last Updated": "",
        }
    )
    report.save("q2_sales_report.xlsx")
"""

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from xlsx_builder import XLSXBuilder, Palette, Fmt

_THIN_BORDER = Border(
    bottom=Side(style="thin", color="DDDDDD")
)


class DataReport(XLSXBuilder):
    """
    General-purpose data table / report.

    The template is column-driven: you define the columns up front and the
    builder handles headers, input rows, totals, and the summary sheet.

    Column spec tuple:  (header: str, width: float, fmt: str, total_func: str | None)
      header      — column label shown in row 1 of the table
      width       — column width in Excel units
      fmt         — number format string (use Fmt constants or custom)
      total_func  — Excel function name for the totals row ("SUM", "AVERAGE",
                    "COUNT", "MAX", "MIN") or None for text/label columns
    """

    # Layout constants
    _MARGIN_COL   = "A"          # empty left margin column
    _MARGIN_WIDTH = 3
    _DATA_START_COL = 2          # 1-indexed; column B
    _TITLE_ROW    = 2
    _META_START   = 4
    _TABLE_HEADER_ROW_OFFSET = 2  # rows after last metadata row before table starts

    def build(self,
              title: str = "Data Report",
              subtitle: str = "",
              columns: list = None,
              data_rows: int = 20,
              metadata: dict = None):
        """
        Build the workbook.

        Args:
            title       : Report title shown on cover and data sheet
            subtitle    : Optional subtitle / description
            columns     : List of (header, width, fmt, total_func) tuples.
                          Falls back to a sensible default if omitted.
            data_rows   : Number of blank input rows in the data table
            metadata    : Dict of label→value pairs shown on the cover sheet

        Returns self for chaining.
        """
        if columns is None:
            columns = self._default_columns()
        if metadata is None:
            metadata = {
                "Prepared By":  "",
                "Department":   "",
                "Period":       "",
                "Last Updated": "",
            }

        self._columns   = columns
        self._data_rows = data_rows
        self._title     = title
        self._subtitle  = subtitle
        self._metadata  = metadata

        self._build_cover()
        self._build_data()
        self._build_summary()

        return self

    # ── Sheet builders ────────────────────────────────────────────────────

    def _build_cover(self):
        ws = self.add_sheet("Cover")
        self.set_col_widths(ws, {
            "A": self._MARGIN_WIDTH,
            "B": 28,
            "C": 40,
            "D": 4,
        })

        # Title
        self.title_row(ws, self._title, "B", "C", self._TITLE_ROW)
        ws.row_dimensions[self._TITLE_ROW].height = 36

        # Subtitle
        if self._subtitle:
            ws.merge_cells("B3:C3")
            c = ws["B3"]
            c.value = self._subtitle
            c.font  = Font(name="Arial", italic=True, size=11, color=Palette.WHITE)
            c.fill  = __import__("openpyxl.styles", fromlist=["PatternFill"]).PatternFill(
                "solid", start_color=Palette.MID_BLUE)
            c.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[3].height = 20

        # Metadata block
        meta_start = self._META_START + (1 if self._subtitle else 0)
        self.section_header_row(ws, "Report Information", "B", "C", meta_start)
        row = meta_start + 1
        for label, value in self._metadata.items():
            self.input_row(ws, label, "B", "C", row, merge_input_to="C")
            if value:
                ws[f"C{row}"].value = value
            row += 1

        # Legend
        self.legend(ws, "B", "C", row + 1)

    def _build_data(self):
        ws = self.add_sheet("Data")

        # Column widths: margin + one per data column
        widths = {self._MARGIN_COL: self._MARGIN_WIDTH}
        for i, (_, width, _, _) in enumerate(self._columns):
            widths[get_column_letter(self._DATA_START_COL + i)] = width
        self.set_col_widths(ws, widths)

        # Title bar
        last_col = get_column_letter(self._DATA_START_COL + len(self._columns) - 1)
        self.title_row(ws, self._title, "B", last_col, 1, size=13)

        # Column headers (row 2)
        header_row = 2
        headers = [col[0] for col in self._columns]
        self.col_header_row(ws, headers, start_col=self._DATA_START_COL,
                            row=header_row, dark=True, height=22)

        # Freeze panes below header, right of margin
        self.freeze_panes(ws, f"{get_column_letter(self._DATA_START_COL)}{header_row + 1}")

        # Data input rows
        first_data_row = header_row + 1
        last_data_row  = first_data_row + self._data_rows - 1

        for r in range(first_data_row, last_data_row + 1):
            alt = (r - first_data_row) % 2 == 1
            for i, (_, _, fmt, _total_func) in enumerate(self._columns):
                col_letter = get_column_letter(self._DATA_START_COL + i)
                cell = ws[f"{col_letter}{r}"]
                self.s.input(cell, fmt=fmt)
                if alt:
                    cell.border = _THIN_BORDER
                ws.row_dimensions[r].height = 18

        # Totals row
        totals_row = last_data_row + 1
        ws.row_dimensions[totals_row].height = 20
        for i, (header, _, fmt, total_func) in enumerate(self._columns):
            col_letter = get_column_letter(self._DATA_START_COL + i)
            cell = ws[f"{col_letter}{totals_row}"]
            if i == 0:
                self.s.section_total(cell, "TOTAL")
            elif total_func:
                data_range = f"{col_letter}{first_data_row}:{col_letter}{last_data_row}"
                self.s.highlight_total(cell, f"={total_func}({data_range})", fmt=fmt)
            else:
                cell.fill = PatternFill("solid", start_color=Palette.SECTION_FILL)

        # Row counter helper (bottom-right, subtle)
        count_col = get_column_letter(self._DATA_START_COL + len(self._columns))
        ws.column_dimensions[count_col].width = 4
        note_row = totals_row + 1
        ws[f"B{note_row}"].value = (
            f'="{self._data_rows} input rows  |  '
            f'Edit column headers in row {header_row} to suit your data"'
        )
        ws[f"B{note_row}"].font = Font(name="Arial", italic=True, size=8,
                                       color="888888")
        ws.merge_cells(f"B{note_row}:{last_col}{note_row}")

    def _build_summary(self):
        ws = self.add_sheet("Summary")

        # Column layout
        widths = {"A": self._MARGIN_WIDTH, "B": 22, "C": 22, "D": 22, "E": 22, "F": 4}
        self.set_col_widths(ws, widths)

        self.title_row(ws, f"{self._title} — Summary", "B", "E", 1, size=13)

        # KPI tiles — auto-generated for SUM/AVERAGE columns
        self.section_header_row(ws, "Key Metrics", "B", "E", 3)
        self.col_header_row(ws, ["Metric", "Value", "Source Column", "Notes"],
                            start_col=2, row=4, dark=False)

        kpi_row = 5
        for i, (header, _, fmt, total_func) in enumerate(self._columns):
            if total_func in ("SUM", "AVERAGE", "MAX", "MIN"):
                col_letter = get_column_letter(self._DATA_START_COL + i)
                self.s.label(ws[f"B{kpi_row}"], header, bold=True)
                # Cross-sheet reference to the totals row on the Data sheet
                data_totals_row = 2 + self._data_rows + 1  # header(1) + data + 1
                self.s.xsheet(
                    ws[f"C{kpi_row}"],
                    f"=Data!{col_letter}{data_totals_row}",
                    fmt=fmt
                )
                self.s.label(ws[f"D{kpi_row}"], f"Data col {col_letter}")
                self.s.input(ws[f"E{kpi_row}"])
                ws.row_dimensions[kpi_row].height = 20
                kpi_row += 1

        # Notes / observations area
        notes_start = kpi_row + 2
        self.section_header_row(ws, "Notes & Observations", "B", "E", notes_start)
        for r in range(notes_start + 1, notes_start + 7):
            ws.merge_cells(f"B{r}:E{r}")
            self.s.textarea(ws[f"B{r}"])
            ws.row_dimensions[r].height = 28

        # Chart placeholder note
        chart_row = notes_start + 9
        self.section_header_row(ws, "Chart Area (insert chart here)", "B", "E", chart_row)
        ws.merge_cells(f"B{chart_row + 1}:E{chart_row + 8}")
        c = ws[f"B{chart_row + 1}"]
        c.value = "← Insert chart referencing Data sheet columns"
        c.font  = Font(name="Arial", italic=True, size=10, color="888888")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill  = PatternFill("solid", start_color=Palette.VERY_LIGHT_BLUE)
        ws.row_dimensions[chart_row + 1].height = 120

    # ── Defaults ──────────────────────────────────────────────────────────

    @staticmethod
    def _default_columns():
        return [
            ("Category",      22, Fmt.TEXT,     None),
            ("Description",   30, Fmt.TEXT,     None),
            ("Quantity",       14, Fmt.INTEGER,  "SUM"),
            ("Unit Price ($)", 16, Fmt.CURRENCY, "AVERAGE"),
            ("Total ($)",      16, Fmt.CURRENCY, "SUM"),
            ("Notes",          24, Fmt.TEXT,     None),
        ]
