"""
xlsx_builder/templates/financial_model.py
------------------------------------------
3-Statement Financial Model template.

Produces a five-sheet workbook:
  1. Cover           — company info, navigation, legend
  2. Assumptions     — all model drivers (revenue, margins, working capital,
                       capex, balance sheet starting values)
  3. Income Statement — P&L with Historical + 5-year projections
  4. Balance Sheet   — Assets / L&E with cash as the CF plug
  5. Cash Flow       — Indirect method; ending cash links back to BS

Statement linkage (no circular references):
  Assumptions → IS  (revenue, margin, tax, interest drivers)
  IS          → BS  (net income → retained earnings; D&A → accumulated depr.)
  IS + BS     → CF  (net income, D&A, working capital changes, capex)
  CF          → BS  (ending cash is the cash plug on BS)

Sign convention (standard financial statement presentation):
  Revenue      : positive
  COGS/OpEx    : negative
  Gross Profit : positive
  EBITDA/EBIT  : positive if profitable
  D&A          : negative on IS; added back (positive) on CF
  Net Income   : positive if profitable

New mechanics vs prior templates:
  - Multi-sheet formula linkage (IS → BS → CF → BS)
  - Two-pass cell writing (beginning cash written after ending cash row exists)
  - Assumption row numbers stored as instance vars for maintainability
  - Row-level _write_row() helper for clean IS/BS/CF construction
  - Balance sheet check row (Assets − L&E = 0)
  - CF ending cash vs BS cash check row

Usage:
    from financial_model import FinancialModel

    model = FinancialModel()
    model.build(
        company_name="Acme Corp",
        currency_symbol="$",
    )
    model.save("acme_financial_model.xlsx")
"""

from openpyxl.styles import PatternFill, Font, Alignment
from xlsx_builder import XLSXBuilder, Palette, Fmt


class FinancialModel(XLSXBuilder):
    """
    3-Statement Financial Model: Income Statement, Balance Sheet, Cash Flow.

    All projection columns (Year 1–5) are formula-driven from the Assumptions
    sheet. The Historical column is entirely user input.

    Instance variables set during build (row number references):
      _a_*          : assumption row numbers on the Assumptions sheet
      _is_*_row     : key row numbers on the Income Statement
      _bs_*_row     : key row numbers on the Balance Sheet
      _cf_*_row     : key row numbers on the Cash Flow
    """

    # ── Column structure ──────────────────────────────────────────────────

    YEARS      = ["Historical", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"]
    COLS       = ["C", "D", "E", "F", "G", "H"]   # one column per year
    PROJ_COLS  = ["D", "E", "F", "G", "H"]         # projection years only
    GROWTH_ASM = None   # set in _build_assumptions: [g2_row, g3_row, g4_row, g5_row]

    _ASM_SHEET = "Assumptions"
    _ASM_COL   = "D"                  # value column on the Assumptions sheet
    _IS_SHEET  = "Income Statement"
    _BS_SHEET  = "Balance Sheet"
    _CF_SHEET  = "Cash Flow"

    # ── Public interface ──────────────────────────────────────────────────

    def build(self,
              company_name: str = "Company Name",
              currency_symbol: str = "$"):
        """
        Build the full 3-statement model workbook.

        Args:
            company_name    : Displayed in title bars across all sheets
            currency_symbol : Prefix for currency headers (e.g. "$", "€", "£")

        Returns self for chaining.
        """
        self._co  = company_name
        self._cur = currency_symbol

        self._build_cover()
        self._build_assumptions()
        self._build_income_statement()
        self._build_balance_sheet()
        self._build_cash_flow()       # also back-fills BS cash and CF beg. cash

        return self

    # ── Private helpers ───────────────────────────────────────────────────

    def _a(self, row_num: int) -> str:
        """Absolute cross-sheet reference to an assumption cell, e.g. Assumptions!$D$6."""
        return f"{self._ASM_SHEET}!${self._ASM_COL}${row_num}"

    def _prior_col(self, col: str) -> str:
        """Return the column letter that precedes col in COLS."""
        return self.COLS[self.COLS.index(col) - 1]

    def _setup_statement_sheet(self, title: str, subtitle: str = "") -> object:
        """Create and configure a standard statement sheet."""
        ws = self.add_sheet(title)
        ws.sheet_view.showGridLines = False
        self.set_col_widths(ws, {
            "A": 3,
            "B": 35,   # row labels
            "C": 14, "D": 14, "E": 14, "F": 14, "G": 14, "H": 14,
            "I": 3,
        })
        self.title_row(ws, f"{self._co}  |  {title}", "B", "H", 1, size=12)
        if subtitle:
            ws.merge_cells("B2:H2")
            c = ws["B2"]
            c.value     = subtitle
            c.font      = Font(name="Arial", italic=True, size=9,
                               color=Palette.WHITE)
            c.fill      = PatternFill("solid", start_color=Palette.MID_BLUE)
            c.alignment = Alignment(horizontal="right", vertical="center",
                                    indent=1)
            ws.row_dimensions[2].height = 14
        return ws

    def _year_header_row(self, ws, row: int):
        """Write year labels as column headers across all data columns."""
        self.col_header_row(ws, self.YEARS, start_col=3, row=row,
                            dark=True, height=20)

    def _write_row(self, ws, row: int, label: str, values: list, fmt: str,
                   style: str = "formula", bold: bool = False,
                   highlight: bool = False, indent: int = 1):
        """
        Write a complete statement row (label + one value per year column).

        Args:
            ws      : target worksheet
            row     : Excel row number
            label   : row label text (column B)
            values  : list of 6 items aligned with self.COLS.
                      Use "" or None to leave a cell blank/unstyled.
            fmt     : number format string
            style   : "formula"  — black formula cell
                      "input"    — blue/yellow input cell
                      "xsheet"   — green cross-sheet reference cell
                      "metric"   — italic derived metric (margins, etc.)
                      "total"    — bold, light-blue highlight formula
            bold    : bold label and value cells
            highlight : apply light-blue fill to formula cells
            indent  : label indent level
        """
        lbl_cell = ws[f"B{row}"]
        lbl_cell.value     = label
        lbl_cell.font      = Font(name="Arial", bold=bold, size=10,
                                  color=Palette.DARK_TEXT)
        lbl_cell.alignment = Alignment(horizontal="left", vertical="center",
                                       indent=indent, wrap_text=True)
        ws.row_dimensions[row].height = 18

        for col, val in zip(self.COLS, values):
            cell = ws[f"{col}{row}"]
            if val in ("", None):
                continue

            if style == "input":
                self.s.input(cell, value=val if not str(val).startswith("=") else "",
                             fmt=fmt)
                if str(val).startswith("="):
                    cell.value = val
            elif style == "xsheet":
                self.s.xsheet(cell, val, fmt=fmt)
                if bold:
                    cell.font = Font(name="Arial", bold=True, size=10,
                                     color=Palette.XSHEET)
            elif style == "metric":
                self.s.italic_metric(cell, val, fmt=fmt)
            elif style == "total":
                self.s.highlight_total(cell, val, fmt=fmt)
            else:  # "formula" (default)
                self.s.formula(cell, val, bold=bold, fmt=fmt)
                if highlight:
                    cell.fill = PatternFill("solid",
                                            start_color=Palette.TOTAL_FILL)

    def _input_historical(self, ws, row: int, fmt: str):
        """Apply input styling to the Historical column (C) of a statement row."""
        self.s.input(ws[f"C{row}"], fmt=fmt)

    def _spacer(self, ws, row: int, height: int = 8):
        ws.row_dimensions[row].height = height

    def _asm_row(self, ws, row: int, label: str, default,
                 fmt: str, note: str = ""):
        """Write a single assumption row (label | input value | optional note)."""
        self.s.label(ws[f"C{row}"], label)
        self.s.input(ws[f"D{row}"], value=default, fmt=fmt)
        if note:
            ws[f"E{row}"].value = note
            ws[f"E{row}"].font  = Font(name="Arial", italic=True, size=9,
                                       color="888888")
            ws[f"E{row}"].alignment = Alignment(horizontal="left",
                                                vertical="center", indent=1)
        ws.row_dimensions[row].height = 18

    # ── Sheet 1: Cover ────────────────────────────────────────────────────

    def _build_cover(self):
        ws = self.add_sheet("Cover")
        ws.sheet_view.showGridLines = False
        self.set_col_widths(ws, {"A": 3, "B": 28, "C": 40, "D": 4})

        self.title_row(ws, self._co, "B", "C", 2, size=20)
        ws.merge_cells("B3:C3")
        c = ws["B3"]
        c.value     = "3-Statement Financial Model"
        c.font      = Font(name="Arial", italic=True, size=12, color=Palette.WHITE)
        c.fill      = PatternFill("solid", start_color=Palette.MID_BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[3].height = 22

        self.section_header_row(ws, "Model Information", "B", "C", 5)
        info = [
            ("Company",          self._co),
            ("Currency",         self._cur),
            ("Projection Period","5 Years"),
            ("Model Date",       ""),
            ("Prepared By",      ""),
            ("Version",          "1.0"),
        ]
        for i, (lbl, val) in enumerate(info, start=6):
            self.input_row(ws, lbl, "B", "C", i, merge_input_to="C")
            if val:
                ws[f"C{i}"].value = val
            ws.row_dimensions[i].height = 20

        self.section_header_row(ws, "Model Sheets", "B", "C", 13)
        nav = [self._ASM_SHEET, self._IS_SHEET, self._BS_SHEET, self._CF_SHEET]
        for i, sheet in enumerate(nav, start=14):
            ws.row_dimensions[i].height = 18
            ws.merge_cells(f"B{i}:C{i}")
            c = ws[f"B{i}"]
            c.value     = f"→   {sheet}"
            c.font      = Font(name="Arial", size=10, color=Palette.MID_BLUE,
                               underline="single")
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        self.legend(ws, "B", "C", 19)

    # ── Sheet 2: Assumptions ──────────────────────────────────────────────

    def _build_assumptions(self):
        ws = self.add_sheet(self._ASM_SHEET)
        ws.sheet_view.showGridLines = False
        self.set_col_widths(ws, {
            "A": 3, "B": 4,
            "C": 36, "D": 18, "E": 38, "F": 3,
        })

        self.title_row(ws, f"{self._co}  |  Assumptions", "C", "E", 1, size=12)

        ws.merge_cells("C3:E3")
        c = ws["C3"]
        c.value     = ("All blue/yellow cells are inputs — change these to "
                       "update the entire model automatically.")
        c.font      = Font(name="Arial", italic=True, size=9, color="555555")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[3].height = 16

        for col, hdr in [("C", "Assumption"), ("D", "Value"), ("E", "Notes")]:
            self.s.col_header(ws[f"{col}4"], hdr, dark=False)
        ws.row_dimensions[4].height = 18

        row = 5

        # ── Revenue ───────────────────────────────────────────────────────
        ws.merge_cells(f"C{row}:E{row}")
        self.section_header_row(ws, "REVENUE", "C", "E", row); row += 1

        self._a_rev_y1 = row
        self._asm_row(ws, row, "Year 1 Revenue", 1_000_000, Fmt.CURRENCY,
                      "Base revenue — the model's starting point"); row += 1
        self._a_g2 = row
        self._asm_row(ws, row, "Year 2 Growth Rate", 0.25, Fmt.PCT); row += 1
        self._a_g3 = row
        self._asm_row(ws, row, "Year 3 Growth Rate", 0.20, Fmt.PCT); row += 1
        self._a_g4 = row
        self._asm_row(ws, row, "Year 4 Growth Rate", 0.15, Fmt.PCT); row += 1
        self._a_g5 = row
        self._asm_row(ws, row, "Year 5 Growth Rate", 0.10, Fmt.PCT); row += 2

        self.GROWTH_ASM = [self._a_g2, self._a_g3, self._a_g4, self._a_g5]

        # ── Income Statement Drivers ──────────────────────────────────────
        ws.merge_cells(f"C{row}:E{row}")
        self.section_header_row(ws, "INCOME STATEMENT DRIVERS", "C", "E", row)
        row += 1

        self._a_cogs = row
        self._asm_row(ws, row, "COGS % of Revenue", 0.40, Fmt.PCT); row += 1
        self._a_rd = row
        self._asm_row(ws, row, "R&D % of Revenue", 0.15, Fmt.PCT); row += 1
        self._a_sm = row
        self._asm_row(ws, row, "Sales & Marketing % of Revenue", 0.20, Fmt.PCT)
        row += 1
        self._a_ga = row
        self._asm_row(ws, row, "G&A % of Revenue", 0.10, Fmt.PCT); row += 1
        self._a_da = row
        self._asm_row(ws, row, "D&A % of Revenue", 0.05, Fmt.PCT,
                      "Depreciation & Amortization as % of revenue"); row += 1
        self._a_int = row
        self._asm_row(ws, row, "Interest Rate on Debt", 0.06, Fmt.PCT,
                      "Applied to the constant debt balance"); row += 1
        self._a_tax = row
        self._asm_row(ws, row, "Tax Rate", 0.25, Fmt.PCT,
                      "Applied only when EBT > 0"); row += 2

        # ── Working Capital ───────────────────────────────────────────────
        ws.merge_cells(f"C{row}:E{row}")
        self.section_header_row(ws, "WORKING CAPITAL (Days)", "C", "E", row)
        row += 1

        self._a_dso = row
        self._asm_row(ws, row, "Days Sales Outstanding (DSO)", 45, Fmt.INTEGER,
                      "AR = Revenue × DSO ÷ 365"); row += 1
        self._a_dio = row
        self._asm_row(ws, row, "Days Inventory Outstanding (DIO)", 30, Fmt.INTEGER,
                      "Inventory = |COGS| × DIO ÷ 365"); row += 1
        self._a_dpo = row
        self._asm_row(ws, row, "Days Payable Outstanding (DPO)", 30, Fmt.INTEGER,
                      "AP = |COGS| × DPO ÷ 365"); row += 2

        # ── Capex & PP&E ──────────────────────────────────────────────────
        ws.merge_cells(f"C{row}:E{row}")
        self.section_header_row(ws, "CAPEX & FIXED ASSETS", "C", "E", row)
        row += 1

        self._a_capex = row
        self._asm_row(ws, row, "Capex % of Revenue", 0.08, Fmt.PCT); row += 1
        self._a_ppe_gross = row
        self._asm_row(ws, row, "Beginning Gross PP&E", 500_000, Fmt.CURRENCY,
                      "Gross PP&E at the start of the Historical period"); row += 1
        self._a_ppe_accum = row
        self._asm_row(ws, row, "Beginning Accumulated Depreciation", -200_000,
                      Fmt.CURRENCY, "Enter as a negative number"); row += 2

        # ── Balance Sheet Starting Values ─────────────────────────────────
        ws.merge_cells(f"C{row}:E{row}")
        self.section_header_row(ws, "BALANCE SHEET — BEGINNING VALUES", "C", "E", row)
        row += 1

        self._a_cash0 = row
        self._asm_row(ws, row, "Beginning Cash", 100_000, Fmt.CURRENCY); row += 1
        self._a_debt = row
        self._asm_row(ws, row, "Total Debt (held constant)", 300_000, Fmt.CURRENCY,
                      "Add a debt schedule for dynamic debt modeling"); row += 1
        self._a_stock = row
        self._asm_row(ws, row, "Common Stock + APIC", 500_000, Fmt.CURRENCY)
        row += 1
        self._a_re0 = row
        self._asm_row(ws, row, "Beginning Retained Earnings", 50_000,
                      Fmt.CURRENCY); row += 2

        # ── Other ─────────────────────────────────────────────────────────
        ws.merge_cells(f"C{row}:E{row}")
        self.section_header_row(ws, "OTHER", "C", "E", row); row += 1

        self._a_div = row
        self._asm_row(ws, row, "Dividend Payout Ratio", 0.00, Fmt.PCT,
                      "% of net income paid as dividends (0 = no dividends)")

    # ── Sheet 3: Income Statement ─────────────────────────────────────────

    def _build_income_statement(self):
        IS = self._IS_SHEET
        ws = self._setup_statement_sheet(
            IS,
            subtitle=("All projection columns (Year 1–5) are formula-driven  |  "
                      "Historical column is editable input")
        )
        self.freeze_panes(ws, "C4")
        self._year_header_row(ws, 3)
        row = 4

        # ── Revenue ───────────────────────────────────────────────────────
        self.section_header_row(ws, "REVENUE", "B", "H", row); row += 1
        self._is_rev_row = row

        rev_vals = [""]                                  # Historical: input
        rev_vals.append(f"={self._a(self._a_rev_y1)}")  # Year 1
        growth_asm = [self._a_g2, self._a_g3, self._a_g4, self._a_g5]
        for i, gr in enumerate(growth_asm):
            prior_col = self.COLS[1 + i]   # D=Y1, E=Y2, F=Y3, G=Y4
            rev_vals.append(f"={prior_col}{row}*(1+{self._a(gr)})")

        self._write_row(ws, row, "Revenue", rev_vals, Fmt.CURRENCY, bold=True)
        self._input_historical(ws, row, Fmt.CURRENCY)
        row += 1

        self._is_cogs_row = row
        cogs_vals = [""] + [
            f"=-{c}{self._is_rev_row}*{self._a(self._a_cogs)}"
            for c in self.PROJ_COLS
        ]
        self._write_row(ws, row, "Cost of Goods Sold (COGS)", cogs_vals, Fmt.CURRENCY)
        self._input_historical(ws, row, Fmt.CURRENCY)
        row += 1

        self._is_gp_row = row
        gp_vals = [""] + [
            f"={c}{self._is_rev_row}+{c}{self._is_cogs_row}"
            for c in self.COLS[1:]
        ]
        self._write_row(ws, row, "Gross Profit", gp_vals, Fmt.CURRENCY,
                        bold=True, highlight=True)
        self._input_historical(ws, row, Fmt.CURRENCY)
        row += 1

        gm_vals = [""] + [
            f"=IFERROR({c}{self._is_gp_row}/{c}{self._is_rev_row},0)"
            for c in self.COLS[1:]
        ]
        self._write_row(ws, row, "  Gross Margin %", gm_vals, Fmt.PCT,
                        style="metric")
        row += 1
        self._spacer(ws, row); row += 1

        # ── Operating Expenses ────────────────────────────────────────────
        self.section_header_row(ws, "OPERATING EXPENSES", "B", "H", row); row += 1

        opex_defs = [
            ("Research & Development",         self._a_rd,  "_is_rd_row"),
            ("Sales & Marketing",               self._a_sm,  "_is_sm_row"),
            ("General & Administrative (G&A)", self._a_ga,  "_is_ga_row"),
        ]
        for label, asm_attr, row_attr in opex_defs:
            setattr(self, row_attr, row)
            vals = [""] + [
                f"=-{c}{self._is_rev_row}*{self._a(asm_attr)}"
                for c in self.PROJ_COLS
            ]
            self._write_row(ws, row, label, vals, Fmt.CURRENCY)
            self._input_historical(ws, row, Fmt.CURRENCY)
            row += 1

        self._is_opex_row = row
        opex_vals = [""] + [
            f"=SUM({c}{self._is_rd_row}:{c}{self._is_ga_row})"
            for c in self.COLS[1:]
        ]
        self._write_row(ws, row, "Total Operating Expenses", opex_vals,
                        Fmt.CURRENCY, bold=True, highlight=True)
        self._input_historical(ws, row, Fmt.CURRENCY)
        row += 1
        self._spacer(ws, row); row += 1

        # ── EBITDA ────────────────────────────────────────────────────────
        self._is_ebitda_row = row
        ebitda_vals = [""] + [
            f"={c}{self._is_gp_row}+{c}{self._is_opex_row}"
            for c in self.COLS[1:]
        ]
        self._write_row(ws, row, "EBITDA", ebitda_vals, Fmt.CURRENCY,
                        bold=True, highlight=True)
        self._input_historical(ws, row, Fmt.CURRENCY)
        row += 1

        em_vals = [""] + [
            f"=IFERROR({c}{self._is_ebitda_row}/{c}{self._is_rev_row},0)"
            for c in self.COLS[1:]
        ]
        self._write_row(ws, row, "  EBITDA Margin %", em_vals, Fmt.PCT,
                        style="metric")
        row += 1

        self._is_da_row = row
        da_vals = [""] + [
            f"=-{c}{self._is_rev_row}*{self._a(self._a_da)}"
            for c in self.PROJ_COLS
        ]
        self._write_row(ws, row, "Depreciation & Amortization (D&A)",
                        da_vals, Fmt.CURRENCY)
        self._input_historical(ws, row, Fmt.CURRENCY)
        row += 1

        self._is_ebit_row = row
        ebit_vals = [""] + [
            f"={c}{self._is_ebitda_row}+{c}{self._is_da_row}"
            for c in self.COLS[1:]
        ]
        self._write_row(ws, row, "EBIT (Operating Income)", ebit_vals,
                        Fmt.CURRENCY, bold=True, highlight=True)
        self._input_historical(ws, row, Fmt.CURRENCY)
        row += 1

        ebit_m_vals = [""] + [
            f"=IFERROR({c}{self._is_ebit_row}/{c}{self._is_rev_row},0)"
            for c in self.COLS[1:]
        ]
        self._write_row(ws, row, "  EBIT Margin %", ebit_m_vals, Fmt.PCT,
                        style="metric")
        row += 1
        self._spacer(ws, row); row += 1

        # ── Below the Line ────────────────────────────────────────────────
        self.section_header_row(ws, "BELOW THE LINE", "B", "H", row); row += 1

        # Interest expense on constant debt balance
        self._is_int_row = row
        int_vals = [""] + [
            f"=ROUND(-{self._a(self._a_debt)}*{self._a(self._a_int)},0)"
            for _ in self.PROJ_COLS
        ]
        self._write_row(ws, row, "Interest Expense", int_vals, Fmt.CURRENCY)
        self._input_historical(ws, row, Fmt.CURRENCY)
        row += 1

        self._is_ebt_row = row
        ebt_vals = [""] + [
            f"={c}{self._is_ebit_row}+{c}{self._is_int_row}"
            for c in self.COLS[1:]
        ]
        self._write_row(ws, row, "Earnings Before Tax (EBT)", ebt_vals,
                        Fmt.CURRENCY, bold=True, highlight=True)
        self._input_historical(ws, row, Fmt.CURRENCY)
        row += 1

        # Tax: only on positive EBT
        self._is_tax_row = row
        tax_vals = [""] + [
            f"=IF({c}{self._is_ebt_row}>0,"
            f"ROUND(-{c}{self._is_ebt_row}*{self._a(self._a_tax)},0),0)"
            for c in self.PROJ_COLS
        ]
        self._write_row(ws, row, "Income Tax Expense", tax_vals, Fmt.CURRENCY)
        self._input_historical(ws, row, Fmt.CURRENCY)
        row += 1

        self._is_ni_row = row
        ni_vals = [""] + [
            f"={c}{self._is_ebt_row}+{c}{self._is_tax_row}"
            for c in self.COLS[1:]
        ]
        self._write_row(ws, row, "Net Income", ni_vals, Fmt.CURRENCY,
                        bold=True, highlight=True)
        self._input_historical(ws, row, Fmt.CURRENCY)
        row += 1

        nm_vals = [""] + [
            f"=IFERROR({c}{self._is_ni_row}/{c}{self._is_rev_row},0)"
            for c in self.COLS[1:]
        ]
        self._write_row(ws, row, "  Net Margin %", nm_vals, Fmt.PCT,
                        style="metric")
        row += 1
        self._spacer(ws, row); row += 1

        # Dividends + retained earnings addition (used by BS)
        self._is_div_row = row
        div_vals = [""] + [
            f"=IF({c}{self._is_ni_row}>0,"
            f"ROUND(-{c}{self._is_ni_row}*{self._a(self._a_div)},0),0)"
            for c in self.PROJ_COLS
        ]
        self._write_row(ws, row, "Dividends", div_vals, Fmt.CURRENCY)
        self._input_historical(ws, row, Fmt.CURRENCY)
        row += 1

        self._is_re_add_row = row
        re_add_vals = [""] + [
            f"={c}{self._is_ni_row}+{c}{self._is_div_row}"
            for c in self.COLS[1:]
        ]
        self._write_row(ws, row, "Addition to Retained Earnings", re_add_vals,
                        Fmt.CURRENCY, bold=True)
        self._input_historical(ws, row, Fmt.CURRENCY)

    # ── Sheet 4: Balance Sheet ────────────────────────────────────────────

    def _build_balance_sheet(self):
        IS = self._IS_SHEET
        ws = self._setup_statement_sheet(
            self._BS_SHEET,
            subtitle=("Cash = Cash Flow ending cash  |  "
                      "Retained Earnings rolls from prior period")
        )
        self.freeze_panes(ws, "C4")
        self._year_header_row(ws, 3)
        self._bs_ws = ws
        row = 4

        # ── ASSETS ───────────────────────────────────────────────────────
        self.section_header_row(ws, "ASSETS", "B", "H", row); row += 1

        # Cash — placeholder; back-filled by _build_cash_flow()
        self._bs_cash_row = row
        self._write_row(ws, row, "Cash & Equivalents", [""] * 6, Fmt.CURRENCY)
        for c in self.COLS:
            self.s.input(ws[f"{c}{row}"], fmt=Fmt.CURRENCY)
        # Historical column: beginning cash from assumptions
        ws[f"C{row}"].value = f"={self._a(self._a_cash0)}"
        ws[f"C{row}"].font  = Font(name="Arial", size=10, color=Palette.XSHEET)
        row += 1

        # Accounts Receivable = Revenue × DSO ÷ 365
        self._bs_ar_row = row
        ar_vals = [""] + [
            f"=ROUND({IS}!{c}{self._is_rev_row}*{self._a(self._a_dso)}/365,0)"
            for c in self.PROJ_COLS
        ]
        self._write_row(ws, row, "Accounts Receivable", ar_vals, Fmt.CURRENCY)
        self._input_historical(ws, row, Fmt.CURRENCY)
        row += 1

        # Inventory = |COGS| × DIO ÷ 365
        self._bs_inv_row = row
        inv_vals = [""] + [
            f"=ROUND(-{IS}!{c}{self._is_cogs_row}*{self._a(self._a_dio)}/365,0)"
            for c in self.PROJ_COLS
        ]
        self._write_row(ws, row, "Inventory", inv_vals, Fmt.CURRENCY)
        self._input_historical(ws, row, Fmt.CURRENCY)
        row += 1

        # Total Current Assets
        self._bs_tca_row = row
        tca_vals = [""] + [
            f"=SUM({c}{self._bs_cash_row}:{c}{self._bs_inv_row})"
            for c in self.COLS[1:]
        ]
        self._write_row(ws, row, "Total Current Assets", tca_vals,
                        Fmt.CURRENCY, style="total", bold=True)
        self._input_historical(ws, row, Fmt.CURRENCY)
        row += 1
        self._spacer(ws, row); row += 1

        # Gross PP&E: Y1 = beginning + capex; Y2+ = prior + capex
        self._bs_ppe_gross_row = row
        ppe_g_vals = [""]
        ppe_g_vals.append(
            f"={self._a(self._a_ppe_gross)}"
            f"+(-{IS}!D{self._is_rev_row}*{self._a(self._a_capex)})"
        )
        for col in ["E", "F", "G", "H"]:
            prior = self._prior_col(col)
            ppe_g_vals.append(
                f"={prior}{row}"
                f"+(-{IS}!{col}{self._is_rev_row}*{self._a(self._a_capex)})"
            )
        self._write_row(ws, row, "Gross PP&E", ppe_g_vals, Fmt.CURRENCY)
        self._input_historical(ws, row, Fmt.CURRENCY)
        row += 1

        # Accumulated Depreciation (negative): Y1 = beginning + D&A; Y2+ = prior + D&A
        self._bs_accum_row = row
        accum_vals = [""]
        accum_vals.append(
            f"={self._a(self._a_ppe_accum)}+{IS}!D{self._is_da_row}"
        )
        for col in ["E", "F", "G", "H"]:
            prior = self._prior_col(col)
            accum_vals.append(f"={prior}{row}+{IS}!{col}{self._is_da_row}")
        self._write_row(ws, row, "Accumulated Depreciation", accum_vals, Fmt.CURRENCY)
        self._input_historical(ws, row, Fmt.CURRENCY)
        row += 1

        # Net PP&E
        self._bs_ppe_net_row = row
        ppe_net_vals = [""] + [
            f"={c}{self._bs_ppe_gross_row}+{c}{self._bs_accum_row}"
            for c in self.COLS[1:]
        ]
        self._write_row(ws, row, "Net PP&E", ppe_net_vals, Fmt.CURRENCY,
                        style="total", bold=True)
        self._input_historical(ws, row, Fmt.CURRENCY)
        row += 1

        # Total Assets
        self._bs_ta_row = row
        ta_vals = [""] + [
            f"={c}{self._bs_tca_row}+{c}{self._bs_ppe_net_row}"
            for c in self.COLS[1:]
        ]
        self._write_row(ws, row, "TOTAL ASSETS", ta_vals, Fmt.CURRENCY,
                        style="total", bold=True)
        self._input_historical(ws, row, Fmt.CURRENCY)
        row += 1
        self._spacer(ws, row); row += 1

        # ── LIABILITIES & EQUITY ─────────────────────────────────────────
        self.section_header_row(ws, "LIABILITIES & EQUITY", "B", "H", row); row += 1

        # Accounts Payable = |COGS| × DPO ÷ 365
        self._bs_ap_row = row
        ap_vals = [""] + [
            f"=ROUND(-{IS}!{c}{self._is_cogs_row}*{self._a(self._a_dpo)}/365,0)"
            for c in self.PROJ_COLS
        ]
        self._write_row(ws, row, "Accounts Payable", ap_vals, Fmt.CURRENCY)
        self._input_historical(ws, row, Fmt.CURRENCY)
        row += 1

        # Total Current Liabilities (= AP for this model)
        self._bs_tcl_row = row
        tcl_vals = [""] + [
            f"={c}{self._bs_ap_row}" for c in self.COLS[1:]
        ]
        self._write_row(ws, row, "Total Current Liabilities", tcl_vals,
                        Fmt.CURRENCY, style="total", bold=True)
        self._input_historical(ws, row, Fmt.CURRENCY)
        row += 1

        # Long-term Debt (constant)
        self._bs_debt_row = row
        debt_vals = [""] + [
            f"={self._a(self._a_debt)}" for _ in self.PROJ_COLS
        ]
        self._write_row(ws, row, "Long-term Debt", debt_vals, Fmt.CURRENCY)
        self._input_historical(ws, row, Fmt.CURRENCY)
        row += 1

        # Total Liabilities
        self._bs_tl_row = row
        tl_vals = [""] + [
            f"={c}{self._bs_tcl_row}+{c}{self._bs_debt_row}"
            for c in self.COLS[1:]
        ]
        self._write_row(ws, row, "Total Liabilities", tl_vals,
                        Fmt.CURRENCY, style="total", bold=True)
        self._input_historical(ws, row, Fmt.CURRENCY)
        row += 1
        self._spacer(ws, row); row += 1

        # Common Stock + APIC (constant)
        self._bs_stock_row = row
        stock_vals = [""] + [
            f"={self._a(self._a_stock)}" for _ in self.PROJ_COLS
        ]
        self._write_row(ws, row, "Common Stock + APIC", stock_vals, Fmt.CURRENCY)
        self._input_historical(ws, row, Fmt.CURRENCY)
        row += 1

        # Retained Earnings: rolls from prior period + IS addition
        self._bs_re_row = row
        re_vals = [""]
        # Year 1: beginning RE + IS addition
        re_vals.append(
            f"={self._a(self._a_re0)}+{IS}!D{self._is_re_add_row}"
        )
        # Year 2–5: prior BS RE + IS addition
        for col in ["E", "F", "G", "H"]:
            prior = self._prior_col(col)
            re_vals.append(
                f"={prior}{row}+{IS}!{col}{self._is_re_add_row}"
            )
        self._write_row(ws, row, "Retained Earnings", re_vals, Fmt.CURRENCY)
        self._input_historical(ws, row, Fmt.CURRENCY)
        row += 1

        # Total Equity
        self._bs_te_row = row
        te_vals = [""] + [
            f"={c}{self._bs_stock_row}+{c}{self._bs_re_row}"
            for c in self.COLS[1:]
        ]
        self._write_row(ws, row, "Total Equity", te_vals,
                        Fmt.CURRENCY, style="total", bold=True)
        self._input_historical(ws, row, Fmt.CURRENCY)
        row += 1

        # Total L&E
        self._bs_tle_row = row
        tle_vals = [""] + [
            f"={c}{self._bs_tl_row}+{c}{self._bs_te_row}"
            for c in self.COLS[1:]
        ]
        self._write_row(ws, row, "TOTAL LIABILITIES & EQUITY", tle_vals,
                        Fmt.CURRENCY, style="total", bold=True)
        self._input_historical(ws, row, Fmt.CURRENCY)
        row += 1
        self._spacer(ws, row, height=12); row += 1

        # Balance check: Assets − L&E = 0
        check_vals = [""] + [
            f"={c}{self._bs_ta_row}-{c}{self._bs_tle_row}"
            for c in self.COLS[1:]
        ]
        self._write_row(ws, row, "CHECK: Assets − L&E  (must = 0)",
                        check_vals, Fmt.CURRENCY, style="metric")
        for col in self.COLS[1:]:
            ws[f"{col}{row}"].font = Font(name="Arial", italic=True, size=10,
                                          color=Palette.XSHEET)
        self._bs_check_row = row

    # ── Sheet 5: Cash Flow ────────────────────────────────────────────────

    def _build_cash_flow(self):
        """
        Build the Cash Flow sheet (indirect method).

        Also performs two back-fills after the ending cash row is known:
          1. BS cash cells (D–H) → CF ending cash
          2. CF beginning cash (Year 2–5) → prior year CF ending cash
        """
        IS = self._IS_SHEET
        BS = self._BS_SHEET
        ws = self._setup_statement_sheet(
            self._CF_SHEET,
            subtitle="Indirect method  |  Ending cash feeds the Balance Sheet cash plug"
        )
        self.freeze_panes(ws, "C4")
        self._year_header_row(ws, 3)
        row = 4

        # ── Operating Activities ──────────────────────────────────────────
        self.section_header_row(ws, "CASH FROM OPERATING ACTIVITIES",
                                "B", "H", row); row += 1

        # Net Income (cross-sheet from IS)
        self._cf_ni_row = row
        ni_vals = [""] + [
            f"={IS}!{c}{self._is_ni_row}" for c in self.PROJ_COLS
        ]
        self._write_row(ws, row, "Net Income", ni_vals, Fmt.CURRENCY,
                        style="xsheet")
        self._input_historical(ws, row, Fmt.CURRENCY)
        row += 1

        # D&A add-back (non-cash; sign flip to make positive)
        self._cf_da_row = row
        da_vals = [""] + [
            f"=-{IS}!{c}{self._is_da_row}" for c in self.PROJ_COLS
        ]
        self._write_row(ws, row, "Add: Depreciation & Amortization",
                        da_vals, Fmt.CURRENCY, style="xsheet")
        self._input_historical(ws, row, Fmt.CURRENCY)
        row += 1

        # Working Capital Changes sub-header
        self.section_header_row(ws, "  Changes in Working Capital",
                                "B", "H", row, height=14); row += 1

        # ΔAR: increase = use of cash (negative)
        self._cf_dar_row = row
        dar_vals = [""]
        for col in self.PROJ_COLS:
            prior = self._prior_col(col)
            dar_vals.append(
                f"=-({BS}!{col}{self._bs_ar_row}-{BS}!{prior}{self._bs_ar_row})"
            )
        self._write_row(ws, row, "  (Increase)/Decrease in Accounts Receivable",
                        dar_vals, Fmt.CURRENCY, indent=2)
        self._input_historical(ws, row, Fmt.CURRENCY)
        row += 1

        # ΔInventory
        self._cf_dinv_row = row
        dinv_vals = [""]
        for col in self.PROJ_COLS:
            prior = self._prior_col(col)
            dinv_vals.append(
                f"=-({BS}!{col}{self._bs_inv_row}-{BS}!{prior}{self._bs_inv_row})"
            )
        self._write_row(ws, row, "  (Increase)/Decrease in Inventory",
                        dinv_vals, Fmt.CURRENCY, indent=2)
        self._input_historical(ws, row, Fmt.CURRENCY)
        row += 1

        # ΔAP: increase = source of cash (positive)
        self._cf_dap_row = row
        dap_vals = [""]
        for col in self.PROJ_COLS:
            prior = self._prior_col(col)
            dap_vals.append(
                f"={BS}!{col}{self._bs_ap_row}-{BS}!{prior}{self._bs_ap_row}"
            )
        self._write_row(ws, row, "  Increase/(Decrease) in Accounts Payable",
                        dap_vals, Fmt.CURRENCY, indent=2)
        self._input_historical(ws, row, Fmt.CURRENCY)
        row += 1

        # Total WC Changes
        self._cf_wc_row = row
        wc_vals = [""] + [
            f"=SUM({c}{self._cf_dar_row}:{c}{self._cf_dap_row})"
            for c in self.COLS[1:]
        ]
        self._write_row(ws, row, "Total Working Capital Changes",
                        wc_vals, Fmt.CURRENCY, bold=True, highlight=True)
        self._input_historical(ws, row, Fmt.CURRENCY)
        row += 1

        # Cash from Operations
        self._cf_cfo_row = row
        cfo_vals = [""] + [
            f"={c}{self._cf_ni_row}+{c}{self._cf_da_row}+{c}{self._cf_wc_row}"
            for c in self.COLS[1:]
        ]
        self._write_row(ws, row, "Cash from Operations (CFO)", cfo_vals,
                        Fmt.CURRENCY, style="total", bold=True)
        self._input_historical(ws, row, Fmt.CURRENCY)
        row += 1
        self._spacer(ws, row); row += 1

        # ── Investing Activities ──────────────────────────────────────────
        self.section_header_row(ws, "CASH FROM INVESTING ACTIVITIES",
                                "B", "H", row); row += 1

        self._cf_capex_row = row
        capex_vals = [""] + [
            f"=-{IS}!{c}{self._is_rev_row}*{self._a(self._a_capex)}"
            for c in self.PROJ_COLS
        ]
        self._write_row(ws, row, "Capital Expenditures (Capex)",
                        capex_vals, Fmt.CURRENCY)
        self._input_historical(ws, row, Fmt.CURRENCY)
        row += 1

        self._cf_cfi_row = row
        cfi_vals = [""] + [
            f"={c}{self._cf_capex_row}" for c in self.COLS[1:]
        ]
        self._write_row(ws, row, "Cash from Investing (CFI)", cfi_vals,
                        Fmt.CURRENCY, style="total", bold=True)
        self._input_historical(ws, row, Fmt.CURRENCY)
        row += 1
        self._spacer(ws, row); row += 1

        # ── Financing Activities ──────────────────────────────────────────
        self.section_header_row(ws, "CASH FROM FINANCING ACTIVITIES",
                                "B", "H", row); row += 1

        # Net debt issuance — input (0 default; model keeps debt constant)
        self._cf_debt_row = row
        debt_vals = ["", 0, 0, 0, 0, 0]
        self._write_row(ws, row, "Net Borrowings / (Repayments)",
                        debt_vals, Fmt.CURRENCY, style="input")
        row += 1

        # Dividends paid (cross-sheet from IS)
        self._cf_div_row = row
        div_vals = [""] + [
            f"={IS}!{c}{self._is_div_row}" for c in self.PROJ_COLS
        ]
        self._write_row(ws, row, "Dividends Paid", div_vals, Fmt.CURRENCY,
                        style="xsheet")
        self._input_historical(ws, row, Fmt.CURRENCY)
        row += 1

        self._cf_cff_row = row
        cff_vals = [""] + [
            f"={c}{self._cf_debt_row}+{c}{self._cf_div_row}"
            for c in self.COLS[1:]
        ]
        self._write_row(ws, row, "Cash from Financing (CFF)", cff_vals,
                        Fmt.CURRENCY, style="total", bold=True)
        self._input_historical(ws, row, Fmt.CURRENCY)
        row += 1
        self._spacer(ws, row, height=12); row += 1

        # ── Cash Reconciliation ───────────────────────────────────────────
        # Net change in cash
        self._cf_net_row = row
        net_vals = [""] + [
            f"={c}{self._cf_cfo_row}+{c}{self._cf_cfi_row}+{c}{self._cf_cff_row}"
            for c in self.COLS[1:]
        ]
        self._write_row(ws, row, "Net Change in Cash", net_vals,
                        Fmt.CURRENCY, bold=True)
        self._input_historical(ws, row, Fmt.CURRENCY)
        row += 1

        # Beginning cash — written as placeholders, fixed up below
        self._cf_beg_row = row
        self._write_row(ws, row, "Beginning Cash", [""] * 6, Fmt.CURRENCY)
        for c in self.COLS:
            self.s.input(ws[f"{c}{row}"], fmt=Fmt.CURRENCY)
        # Historical and Year 1 both reference beginning cash from assumptions
        for c in ["C", "D"]:
            ws[f"{c}{row}"].value = f"={self._a(self._a_cash0)}"
            ws[f"{c}{row}"].font  = Font(name="Arial", size=10,
                                         color=Palette.XSHEET)
        row += 1

        # Ending cash
        self._cf_end_row = row
        end_vals = [""] + [
            f"={c}{self._cf_net_row}+{c}{self._cf_beg_row}"
            for c in self.COLS[1:]
        ]
        self._write_row(ws, row, "Ending Cash", end_vals, Fmt.CURRENCY,
                        style="total", bold=True)
        self._input_historical(ws, row, Fmt.CURRENCY)
        row += 1

        # ── Two-pass back-fills ───────────────────────────────────────────

        # 1. CF beginning cash Year 2–5 → prior year CF ending cash
        for col in ["E", "F", "G", "H"]:
            prior = self._prior_col(col)
            cell = ws[f"{col}{self._cf_beg_row}"]
            cell.value          = f"={prior}{self._cf_end_row}"
            cell.font           = Font(name="Arial", size=10, color=Palette.FORMULA)
            cell.number_format  = Fmt.CURRENCY
            cell.fill = PatternFill(fill_type=None)  # clear input yellow

        # 2. BS cash cells (Year 1–5) → CF ending cash on this sheet
        bs = self._bs_ws
        for col in self.PROJ_COLS:
            cell = bs[f"{col}{self._bs_cash_row}"]
            cell.value          = f"='{self._CF_SHEET}'!{col}{self._cf_end_row}"
            cell.font           = Font(name="Arial", size=10, color=Palette.XSHEET)
            cell.number_format  = Fmt.CURRENCY
            cell.fill = PatternFill(fill_type=None)  # not user input — remove yellow

        # CF ending cash vs BS cash check
        check_vals = [""] + [
            f"={c}{self._cf_end_row}-'{self._BS_SHEET}'!{c}{self._bs_cash_row}"
            for c in self.PROJ_COLS
        ]
        self._write_row(ws, row,
                        "CHECK: CF Ending Cash − BS Cash  (must = 0)",
                        check_vals, Fmt.CURRENCY, style="metric")
        for col in self.PROJ_COLS:
            ws[f"{col}{row}"].font = Font(name="Arial", italic=True, size=10,
                                          color=Palette.XSHEET)
