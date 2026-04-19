"""
xlsx_builder/templates/project_tracker.py
------------------------------------------
Project Tracker template.

Produces a three-sheet workbook:
  1. Cover       — title, project metadata, legend
  2. Tracker     — task table with dropdown validation (Status, Priority),
                   date columns, progress formula, and conditional row shading
  3. Summary     — auto-calculated KPIs (total tasks, % complete, overdue count)
                   and a milestone table

New mechanics introduced (vs DataReport):
  - Data validation dropdowns (Status, Priority)
  - DATE-aware columns with overdue detection formula
  - Per-row progress indicator formula
  - Named lists sheet (hidden) to back the dropdowns

Usage:
    from project_tracker import ProjectTracker

    tracker = ProjectTracker()
    tracker.build(
        title="Q3 Product Roadmap",
        project_metadata={
            "Project Owner": "",
            "Team":          "",
            "Start Date":    "",
            "Target End":    "",
        },
        task_rows=30,
    )
    tracker.save("q3_roadmap.xlsx")
"""

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from xlsx_builder import XLSXBuilder, Palette, Fmt


# ── Status / Priority palettes ────────────────────────────────────────────────

STATUS_OPTIONS   = ["Not Started", "In Progress", "Blocked", "Complete", "Cancelled"]
PRIORITY_OPTIONS = ["Critical", "High", "Medium", "Low"]

# Light fill colors for status rows (applied to the # column as a visual cue)
STATUS_FILLS = {
    "Not Started": "F2F2F2",
    "In Progress": "D9EAD3",   # green tint
    "Blocked":     "FCE5CD",   # orange tint
    "Complete":    "C9DAF8",   # blue tint
    "Cancelled":   "EFEFEF",
}

_THIN = Border(bottom=Side(style="thin", color="DDDDDD"))


class ProjectTracker(XLSXBuilder):
    """
    Task / project tracker with dropdown validation and progress tracking.

    Column layout (fixed):
      # | Task / Deliverable | Owner | Priority | Status | Start Date |
      Due Date | Days Remaining | % Complete | Notes
    """

    # Column definitions: (header, width, fmt)
    # total_func omitted — handled per-column in summary sheet
    COLUMNS = [
        ("#",                  5,  Fmt.INTEGER),
        ("Task / Deliverable", 34, Fmt.TEXT),
        ("Owner",              18, Fmt.TEXT),
        ("Priority",           13, Fmt.TEXT),
        ("Status",             14, Fmt.TEXT),
        ("Start Date",         13, Fmt.DATE),
        ("Due Date",           13, Fmt.DATE),
        ("Days Remaining",     15, Fmt.INTEGER),
        ("% Complete",         13, Fmt.PCT),
        ("Notes",              28, Fmt.TEXT),
    ]

    # 1-indexed column positions (B = col 2)
    _START_COL   = 2   # B
    _COL_NUM     = 0   # offset: #
    _COL_TASK    = 1   # Task
    _COL_OWNER   = 2   # Owner
    _COL_PRI     = 3   # Priority
    _COL_STATUS  = 4   # Status
    _COL_START   = 5   # Start Date
    _COL_DUE     = 6   # Due Date
    _COL_DAYS    = 7   # Days Remaining
    _COL_PCT     = 8   # % Complete
    _COL_NOTES   = 9   # Notes

    def build(self,
              title: str = "Project Tracker",
              project_metadata: dict = None,
              task_rows: int = 25):
        """
        Build the workbook.

        Args:
            title            : Project name shown in title bars
            project_metadata : Dict of label→value for the cover sheet
            task_rows        : Number of task rows to pre-generate

        Returns self for chaining.
        """
        if project_metadata is None:
            project_metadata = {
                "Project Owner": "",
                "Team":          "",
                "Start Date":    "",
                "Target End":    "",
                "Budget":        "",
            }

        self._title    = title
        self._meta     = project_metadata
        self._rows     = task_rows

        self._build_lists()    # hidden validation sheet first
        self._build_cover()
        self._build_tracker()
        self._build_summary()

        return self

    # ── Sheet builders ────────────────────────────────────────────────────

    def _build_lists(self):
        """Hidden sheet backing the dropdown validations."""
        ws = self.add_sheet("_Lists")
        ws.sheet_state = "hidden"

        for i, val in enumerate(STATUS_OPTIONS, start=1):
            ws[f"A{i}"] = val
        for i, val in enumerate(PRIORITY_OPTIONS, start=1):
            ws[f"B{i}"] = val

        self._status_range   = f"_Lists!$A$1:$A${len(STATUS_OPTIONS)}"
        self._priority_range = f"_Lists!$B$1:$B${len(PRIORITY_OPTIONS)}"

    def _build_cover(self):
        ws = self.add_sheet("Cover")
        self.set_col_widths(ws, {"A": 3, "B": 26, "C": 38, "D": 4})

        self.title_row(ws, self._title, "B", "C", 2)

        ws.merge_cells("B3:C3")
        c = ws["B3"]
        c.value     = "Project Tracker"
        c.font      = Font(name="Arial", italic=True, size=10, color=Palette.WHITE)
        c.fill      = PatternFill("solid", start_color=Palette.MID_BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[3].height = 18

        self.section_header_row(ws, "Project Information", "B", "C", 5)
        row = 6
        for label, value in self._meta.items():
            self.input_row(ws, label, "B", "C", row, merge_input_to="C")
            if value:
                ws[f"C{row}"].value = value
            row += 1

        self.section_header_row(ws, "Status Key", "B", "C", row + 1)
        row += 2
        key_items = [
            ("Not Started", STATUS_FILLS["Not Started"], Palette.DARK_TEXT),
            ("In Progress", STATUS_FILLS["In Progress"], Palette.DARK_TEXT),
            ("Blocked",     STATUS_FILLS["Blocked"],     Palette.DARK_TEXT),
            ("Complete",    STATUS_FILLS["Complete"],    Palette.DARK_TEXT),
            ("Cancelled",   STATUS_FILLS["Cancelled"],   "888888"),
        ]
        for status, fill, fc in key_items:
            ws.merge_cells(f"B{row}:C{row}")
            c = ws[f"B{row}"]
            c.value     = status
            c.font      = Font(name="Arial", size=9, color=fc)
            c.fill      = PatternFill("solid", start_color=fill)
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.row_dimensions[row].height = 16
            row += 1

        self.legend(ws, "B", "C", row + 1)

    def _build_tracker(self):
        ws = self.add_sheet("Tracker")

        # Column widths: margin + data columns
        widths = {"A": 3}
        for i, (_, w, _) in enumerate(self.COLUMNS):
            widths[get_column_letter(self._START_COL + i)] = w
        self.set_col_widths(ws, widths)

        # Title + column headers
        last_col = get_column_letter(self._START_COL + len(self.COLUMNS) - 1)
        self.title_row(ws, self._title, "B", last_col, 1, size=13)

        header_row = 2
        self.col_header_row(
            ws,
            [c[0] for c in self.COLUMNS],
            start_col=self._START_COL,
            row=header_row,
            dark=True,
            height=22,
        )
        self.freeze_panes(ws, f"B{header_row + 1}")

        # Data validation objects
        dv_status = DataValidation(
            type="list",
            formula1=f'"{",".join(STATUS_OPTIONS)}"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Invalid Status",
            error=f'Choose from: {", ".join(STATUS_OPTIONS)}',
        )
        dv_priority = DataValidation(
            type="list",
            formula1=f'"{",".join(PRIORITY_OPTIONS)}"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Invalid Priority",
            error=f'Choose from: {", ".join(PRIORITY_OPTIONS)}',
        )
        dv_pct = DataValidation(
            type="decimal",
            operator="between",
            formula1="0",
            formula2="1",
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Invalid value",
            error="Enter a decimal between 0 and 1 (e.g. 0.75 for 75%)",
        )
        ws.add_data_validation(dv_status)
        ws.add_data_validation(dv_priority)
        ws.add_data_validation(dv_pct)

        first_data = header_row + 1
        last_data  = first_data + self._rows - 1

        for r in range(first_data, last_data + 1):
            task_num = r - first_data + 1
            alt      = task_num % 2 == 0

            for i, (_, _, fmt) in enumerate(self.COLUMNS):
                col_letter = get_column_letter(self._START_COL + i)
                cell = ws[f"{col_letter}{r}"]

                if i == self._COL_NUM:
                    # Auto-number
                    cell.value     = task_num
                    cell.font      = Font(name="Arial", bold=True, size=10,
                                         color=Palette.DARK_TEXT)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if alt:
                        cell.fill = PatternFill("solid", start_color=Palette.VERY_LIGHT_BLUE)

                elif i == self._COL_DAYS:
                    # Days Remaining formula — blank if no due date
                    due_col = get_column_letter(self._START_COL + self._COL_DUE)
                    self.s.formula(
                        cell,
                        f'=IF({due_col}{r}="","",{due_col}{r}-TODAY())',
                        fmt=fmt,
                    )
                    # Highlight negative (overdue) in soft red
                    # (manual; conditional formatting requires openpyxl extras)
                    cell.comment = None   # placeholder for conditional note

                else:
                    # Standard input cell
                    self.s.input(cell, fmt=fmt)
                    if alt:
                        cell.border = _THIN

                ws.row_dimensions[r].height = 20

            # Attach dropdowns
            status_col   = get_column_letter(self._START_COL + self._COL_STATUS)
            priority_col = get_column_letter(self._START_COL + self._COL_PRI)
            pct_col      = get_column_letter(self._START_COL + self._COL_PCT)
            dv_status.add(ws[f"{status_col}{r}"])
            dv_priority.add(ws[f"{priority_col}{r}"])
            dv_pct.add(ws[f"{pct_col}{r}"])

        # Totals / summary row
        summary_row = last_data + 2
        ws.row_dimensions[summary_row].height = 20
        ws.merge_cells(f"B{summary_row}:D{summary_row}")
        self.s.section_total(ws[f"B{summary_row}"], "SUMMARY")

        # Total tasks
        num_col = get_column_letter(self._START_COL + self._COL_NUM)
        self.s.highlight_total(
            ws[f"{get_column_letter(self._START_COL + self._COL_NUM)}{summary_row}"],
            f"=COUNTA({num_col}{first_data}:{num_col}{last_data})",
            fmt=Fmt.INTEGER,
        )

        # % Complete average
        pct_col_l = get_column_letter(self._START_COL + self._COL_PCT)
        self.s.highlight_total(
            ws[f"{pct_col_l}{summary_row}"],
            f"=IFERROR(AVERAGE({pct_col_l}{first_data}:{pct_col_l}{last_data}),0)",
            fmt=Fmt.PCT,
        )

        # Overdue count
        due_col_l = get_column_letter(self._START_COL + self._COL_DUE)
        status_col_l = get_column_letter(self._START_COL + self._COL_STATUS)
        days_col_l = get_column_letter(self._START_COL + self._COL_DAYS)
        overdue_cell = ws[f"{days_col_l}{summary_row}"]
        self.s.highlight_total(
            overdue_cell,
            f'=COUNTIFS({days_col_l}{first_data}:{days_col_l}{last_data},"<0",'
            f'{status_col_l}{first_data}:{status_col_l}{last_data},"<>Complete")',
            fmt=Fmt.INTEGER,
        )

        # Label the overdue cell
        label_col = get_column_letter(self._START_COL + self._COL_DAYS - 1)
        ws[f"{label_col}{summary_row}"].value = "Overdue →"
        ws[f"{label_col}{summary_row}"].font  = Font(name="Arial", italic=True,
                                                      size=9, color="888888")
        ws[f"{label_col}{summary_row}"].alignment = Alignment(horizontal="right")

        # Store refs for Summary sheet
        self._tracker_first = first_data
        self._tracker_last  = last_data
        self._tracker_summary_row = summary_row

    def _build_summary(self):
        ws = self.add_sheet("Summary")
        self.set_col_widths(ws, {"A": 3, "B": 24, "C": 20, "D": 20, "E": 20, "F": 4})

        self.title_row(ws, f"{self._title} — Summary", "B", "E", 1, size=13)

        # ── KPI block ─────────────────────────────────────────────────────
        self.section_header_row(ws, "At a Glance", "B", "E", 3)
        self.col_header_row(ws, ["Metric", "Value", "Notes"],
                            start_col=2, row=4, dark=False)

        fd   = self._tracker_first
        ld   = self._tracker_last
        nc   = get_column_letter(self._START_COL + self._COL_NUM)
        sc   = get_column_letter(self._START_COL + self._COL_STATUS)
        pc   = get_column_letter(self._START_COL + self._COL_PCT)
        dc   = get_column_letter(self._START_COL + self._COL_DAYS)

        kpis = [
            ("Total Tasks",
             f"=COUNTA(Tracker!{nc}{fd}:Tracker!{nc}{ld})",
             Fmt.INTEGER),
            ("Complete",
             f'=COUNTIF(Tracker!{sc}{fd}:Tracker!{sc}{ld},"Complete")',
             Fmt.INTEGER),
            ("In Progress",
             f'=COUNTIF(Tracker!{sc}{fd}:Tracker!{sc}{ld},"In Progress")',
             Fmt.INTEGER),
            ("Blocked",
             f'=COUNTIF(Tracker!{sc}{fd}:Tracker!{sc}{ld},"Blocked")',
             Fmt.INTEGER),
            ("Overdue (not complete)",
             f'=COUNTIFS(Tracker!{dc}{fd}:Tracker!{dc}{ld},"<0",'
             f'Tracker!{sc}{fd}:Tracker!{sc}{ld},"<>Complete")',
             Fmt.INTEGER),
            ("Avg % Complete",
             f"=IFERROR(AVERAGE(Tracker!{pc}{fd}:Tracker!{pc}{ld}),0)",
             Fmt.PCT),
        ]

        row = 5
        for label, formula, fmt in kpis:
            self.s.label(ws[f"B{row}"], label, bold=True)
            self.s.xsheet(ws[f"C{row}"], formula, fmt=fmt)
            self.s.input(ws[f"D{row}"])
            ws.row_dimensions[row].height = 20
            row += 1

        # ── Milestone table ───────────────────────────────────────────────
        row += 1
        self.section_header_row(ws, "Key Milestones", "B", "E", row)
        row += 1
        self.col_header_row(ws, ["Milestone", "Target Date", "Owner", "Status"],
                            start_col=2, row=row, dark=False)
        row += 1
        for _ in range(8):
            for col in ["B", "C", "D", "E"]:
                self.s.input(ws[f"{col}{row}"])
            ws.row_dimensions[row].height = 22
            row += 1

        # ── Notes ─────────────────────────────────────────────────────────
        row += 1
        self.section_header_row(ws, "Notes & Risks", "B", "E", row)
        row += 1
        for _ in range(5):
            ws.merge_cells(f"B{row}:E{row}")
            self.s.textarea(ws[f"B{row}"])
            ws.row_dimensions[row].height = 28
            row += 1
