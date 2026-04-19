"""
xlsx_builder/base.py
--------------------
Base classes for generating professional Excel workbooks.

Color conventions (industry standard):
  Blue text   (#0000FF) + Yellow fill — user inputs / hardcoded assumptions
  Black text  (#000000)               — formulas and calculated values
  Green text  (#008000)               — cross-sheet references
  Red text    (#FF0000)               — links to external files
  Light blue fill                     — calculated totals / highlights
  Mid-blue fill (white text)          — sub-headers
  Dark navy fill (white text)         — title blocks and main headers

Number formats:
  Currency    $#,##0;($#,##0);"-"
  Percentage  0.0%;(0.0%);"-"
  Multiple    0.0x
  Integer     #,##0;(#,##0);"-"
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Palette ──────────────────────────────────────────────────────────────────

class Palette:
    DARK_NAVY       = "1F3864"
    MID_BLUE        = "2E5FA3"
    LIGHT_BLUE      = "BDD7EE"
    VERY_LIGHT_BLUE = "DEEAF1"
    WHITE           = "FFFFFF"
    DARK_TEXT       = "1A1A1A"
    LIGHT_GRAY      = "F2F2F2"
    # Text colors (convention)
    INPUT           = "0000FF"   # blue  — user inputs
    FORMULA         = "1A1A1A"   # black — calculated
    XSHEET          = "008000"   # green — cross-sheet
    EXTERNAL        = "FF0000"   # red   — external file links
    # Fill colors (convention)
    INPUT_FILL      = "FFFF00"   # yellow — input cells
    TOTAL_FILL      = "BDD7EE"   # light blue — totals / highlights
    SECTION_FILL    = "DEEAF1"   # very light blue — section rows


# ── Number formats ────────────────────────────────────────────────────────────

class Fmt:
    CURRENCY    = '$#,##0;($#,##0);"-"'
    CURRENCY_MM = '$#,##0.0;($#,##0.0);"-"'   # millions
    PCT         = '0.0%;(0.0%);"-"'
    PCT_1       = '0%;(0%);"-"'
    MULTIPLE    = '0.0x'
    INTEGER     = '#,##0;(#,##0);"-"'
    DATE        = 'MM/DD/YYYY'
    TEXT        = '@'


# ── Borders ───────────────────────────────────────────────────────────────────

class Borders:
    _thin  = Side(style="thin",   color="CCCCCC")
    _thick = Side(style="medium", color="1F3864")

    @classmethod
    def bottom_thin(cls):
        return Border(bottom=cls._thin)

    @classmethod
    def all_thin(cls):
        return Border(top=cls._thin, bottom=cls._thin,
                      left=cls._thin, right=cls._thin)

    @classmethod
    def bottom_thick(cls):
        return Border(bottom=cls._thick)


# ── StyleKit ─────────────────────────────────────────────────────────────────

class StyleKit:
    """
    Applies consistent cell styles across all templates.
    All methods accept a cell and mutate it in place.
    """

    # ── Title / header styles ─────────────────────────────────────────────

    def title(self, cell, text, size=16):
        """Large dark-navy title block (white text)."""
        cell.value = text
        cell.font      = Font(name="Arial", bold=True, size=size, color=Palette.WHITE)
        cell.fill      = PatternFill("solid", start_color=Palette.DARK_NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def section_header(self, cell, text):
        """Mid-blue sub-header (white text), left-aligned."""
        cell.value = text
        cell.font      = Font(name="Arial", bold=True, size=10, color=Palette.WHITE)
        cell.fill      = PatternFill("solid", start_color=Palette.MID_BLUE)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    def col_header(self, cell, text, dark=False):
        """Column header — dark navy or light blue background."""
        cell.value = text
        if dark:
            cell.font = Font(name="Arial", bold=True, size=10, color=Palette.WHITE)
            cell.fill = PatternFill("solid", start_color=Palette.DARK_NAVY)
        else:
            cell.font = Font(name="Arial", bold=True, size=10, color=Palette.DARK_TEXT)
            cell.fill = PatternFill("solid", start_color=Palette.LIGHT_BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── Data cell styles ──────────────────────────────────────────────────

    def label(self, cell, text, bold=False, indent=1):
        """Plain row label (black text, no fill)."""
        cell.value = text
        cell.font      = Font(name="Arial", bold=bold, size=10, color=Palette.DARK_TEXT)
        cell.alignment = Alignment(horizontal="left", vertical="center",
                                   indent=indent, wrap_text=True)

    def input(self, cell, value="", fmt=None):
        """
        User-editable input cell.
        Blue text on yellow fill. Optionally apply a number format.
        """
        cell.value = value
        cell.font      = Font(name="Arial", size=10, color=Palette.INPUT)
        cell.fill      = PatternFill("solid", start_color=Palette.INPUT_FILL)
        cell.alignment = Alignment(horizontal="left", vertical="top",
                                   indent=1, wrap_text=True)
        if fmt:
            cell.number_format = fmt

    def formula(self, cell, expression, bold=False, fmt=None, highlight=False):
        """
        Formula cell. Black text; optionally light-blue highlighted (for totals).
        expression should be a full Excel formula string, e.g. '=SUM(B2:B10)'.
        """
        cell.value = expression
        cell.font  = Font(name="Arial", bold=bold, size=10, color=Palette.FORMULA)
        cell.alignment = Alignment(horizontal="right", vertical="center")
        if highlight:
            cell.fill = PatternFill("solid", start_color=Palette.TOTAL_FILL)
        if fmt:
            cell.number_format = fmt

    def xsheet(self, cell, expression, fmt=None):
        """
        Cross-sheet reference cell (green text).
        expression is a full Excel formula, e.g. '=Assumptions!B5'.
        """
        cell.value = expression
        cell.font  = Font(name="Arial", size=10, color=Palette.XSHEET)
        cell.alignment = Alignment(horizontal="right", vertical="center")
        if fmt:
            cell.number_format = fmt

    def section_total(self, cell, text_or_formula, fmt=None):
        """Section total row — very light blue fill, bold black text."""
        cell.value = text_or_formula
        cell.font  = Font(name="Arial", bold=True, size=10, color=Palette.DARK_TEXT)
        cell.fill  = PatternFill("solid", start_color=Palette.SECTION_FILL)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        if fmt:
            cell.number_format = fmt

    def highlight_total(self, cell, expression, fmt=None):
        """Key total — light blue fill, bold, right-aligned formula."""
        cell.value = expression
        cell.font  = Font(name="Arial", bold=True, size=10, color=Palette.DARK_TEXT)
        cell.fill  = PatternFill("solid", start_color=Palette.TOTAL_FILL)
        cell.alignment = Alignment(horizontal="right", vertical="center")
        if fmt:
            cell.number_format = fmt

    def italic_metric(self, cell, expression, fmt=None):
        """Italic derived metric (e.g. margin %) — no fill."""
        cell.value = expression
        cell.font  = Font(name="Arial", italic=True, size=10, color=Palette.DARK_TEXT)
        cell.alignment = Alignment(horizontal="right", vertical="center")
        if fmt:
            cell.number_format = fmt

    def textarea(self, cell, value=""):
        """Multi-line input area — yellow fill, top-left aligned, wraps."""
        cell.value = value
        cell.font      = Font(name="Arial", size=10, color=Palette.INPUT)
        cell.fill      = PatternFill("solid", start_color=Palette.INPUT_FILL)
        cell.alignment = Alignment(horizontal="left", vertical="top",
                                   indent=1, wrap_text=True)


# ── XLSXBuilder ───────────────────────────────────────────────────────────────

class XLSXBuilder:
    """
    Base class for all template builders.

    Subclasses should:
      1. Call super().__init__()
      2. Implement build() to construct the workbook
      3. Call self.save(path) to write the file

    Attributes:
      wb (Workbook)  : the openpyxl workbook
      s  (StyleKit)  : shared style helper
    """

    DEFAULT_FONT = "Arial"
    DEFAULT_ROW_HEIGHT = 18
    DEFAULT_HEADER_HEIGHT = 30

    def __init__(self):
        self.wb = Workbook()
        self.wb.remove(self.wb.active)   # start with no sheets; subclass adds its own
        self.s = StyleKit()

    # ── Workbook-level helpers ────────────────────────────────────────────

    def add_sheet(self, title, gridlines=False):
        """Add a new worksheet and return it."""
        ws = self.wb.create_sheet(title)
        ws.sheet_view.showGridLines = gridlines
        return ws

    def save(self, path):
        """Save the workbook to disk."""
        self.wb.save(path)
        return path

    # ── Sheet-level helpers ───────────────────────────────────────────────

    def set_col_widths(self, ws, widths: dict):
        """
        Set column widths from a dict mapping column letter to width.
        e.g. {"A": 4, "B": 28, "C": 40}
        """
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

    def set_row_height(self, ws, row, height):
        ws.row_dimensions[row].height = height

    def freeze_panes(self, ws, cell="B2"):
        """Freeze rows above and columns left of the given cell."""
        ws.freeze_panes = cell

    # ── Merge + style shortcuts ───────────────────────────────────────────

    def title_row(self, ws, text, col_start, col_end, row, size=16):
        """
        Merge cells across col_start:col_end at row, apply title style.
        Returns the cell.
        """
        ref = f"{col_start}{row}:{col_end}{row}"
        ws.merge_cells(ref)
        cell = ws[f"{col_start}{row}"]
        self.s.title(cell, text, size=size)
        ws.row_dimensions[row].height = self.DEFAULT_HEADER_HEIGHT
        return cell

    def section_header_row(self, ws, text, col_start, col_end, row, height=18):
        """Merge and apply section_header style."""
        ws.merge_cells(f"{col_start}{row}:{col_end}{row}")
        cell = ws[f"{col_start}{row}"]
        self.s.section_header(cell, text)
        ws.row_dimensions[row].height = height
        return cell

    def col_header_row(self, ws, headers: list, start_col: int, row: int,
                       dark=False, height=18):
        """
        Write a row of column headers.
        headers: list of strings
        start_col: 1-indexed column number
        """
        for i, text in enumerate(headers):
            col_letter = get_column_letter(start_col + i)
            self.s.col_header(ws[f"{col_letter}{row}"], text, dark=dark)
        ws.row_dimensions[row].height = height

    def input_row(self, ws, label, label_col, input_col, row,
                  merge_input_to=None, height=22, fmt=None):
        """
        Write a label + input cell pair (or merged input range).
        merge_input_to: end column letter if input should be merged across columns.
        """
        self.s.label(ws[f"{label_col}{row}"], label, bold=True)
        if merge_input_to:
            ws.merge_cells(f"{input_col}{row}:{merge_input_to}{row}")
        self.s.input(ws[f"{input_col}{row}"], fmt=fmt)
        ws.row_dimensions[row].height = height

    def assumptions_block(self, ws, assumptions: list, label_col, value_col, start_row,
                          section_title=None, title_end_col=None):
        """
        Render a block of assumption rows.

        assumptions: list of (label, default_value, fmt) tuples
        Returns the next available row after the block.

        Example:
            builder.assumptions_block(ws, [
                ("Revenue Growth Rate", 0.30, Fmt.PCT),
                ("COGS as % of Revenue", 0.40, Fmt.PCT),
            ], label_col="B", value_col="C", start_row=5)
        """
        row = start_row

        if section_title:
            end_col = title_end_col or value_col
            self.section_header_row(ws, section_title, label_col, end_col, row)
            row += 1

        for label, default, fmt in assumptions:
            self.s.label(ws[f"{label_col}{row}"], label)
            cell = ws[f"{value_col}{row}"]
            self.s.input(cell, value=default, fmt=fmt)
            ws.row_dimensions[row].height = self.DEFAULT_ROW_HEIGHT
            row += 1

        return row

    def data_rows(self, ws, rows: list, col_start: int, start_row: int,
                  row_height=20):
        """
        Write plain data rows (no special styling).
        rows: list of lists — each inner list is one row of values.
        col_start: 1-indexed column number.
        Returns the next available row.
        """
        for r_data in rows:
            col = col_start
            for value in r_data:
                ws.cell(row=start_row, column=col, value=value)
                col += 1
            ws.row_dimensions[start_row].height = row_height
            start_row += 1
        return start_row

    def legend(self, ws, col_start, col_end, row):
        """
        Render a color-coding legend block.
        Useful on cover sheets.
        """
        ws.merge_cells(f"{col_start}{row}:{col_end}{row}")
        ws[f"{col_start}{row}"].value = "Color Legend"
        ws[f"{col_start}{row}"].font = Font(name="Arial", bold=True, size=9)
        row += 1
        items = [
            ("Blue text / yellow fill = User input or assumption",
             Palette.INPUT, Palette.INPUT_FILL),
            ("Black text = Formula / auto-calculated value",
             Palette.FORMULA, Palette.WHITE),
            ("Green text = Cross-sheet reference",
             Palette.XSHEET, Palette.WHITE),
        ]
        for text, fc, bg in items:
            ws.merge_cells(f"{col_start}{row}:{col_end}{row}")
            c = ws[f"{col_start}{row}"]
            c.value = text
            c.font  = Font(name="Arial", size=9, color=fc)
            c.fill  = PatternFill("solid", start_color=bg)
            ws.row_dimensions[row].height = 16
            row += 1
        return row

    # ── Utility ───────────────────────────────────────────────────────────

    def col(self, n: int) -> str:
        """Return the Excel column letter for a 1-indexed column number."""
        return get_column_letter(n)

    def ref(self, col_letter: str, row: int) -> str:
        """Convenience: return a cell reference string, e.g. 'B12'."""
        return f"{col_letter}{row}"

    def build(self):
        """Override in subclasses to construct the workbook."""
        raise NotImplementedError("Subclasses must implement build()")
