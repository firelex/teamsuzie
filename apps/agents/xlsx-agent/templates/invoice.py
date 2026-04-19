"""
xlsx_builder/templates/invoice.py
----------------------------------
Invoice template.

Produces a single polished invoice sheet — a fixed-layout form rather than
a row-driven table. Architecturally distinct from DataReport and ProjectTracker:
the structure is positional, not iterative.

Sheet layout:
  - Header band     : company logo placeholder, company info, "INVOICE" title
  - Address block   : Bill To / Ship To side-by-side
  - Invoice meta    : Invoice #, Date, Due Date, PO #
  - Line items      : Description, Qty, Unit Price, Amount (formula)
  - Totals block    : Subtotal, Discount, Tax, Total Due
  - Footer          : Payment terms, bank details, notes

New mechanics introduced (vs Tracker):
  - Fixed positional layout (not row-driven)
  - Mixed-width column grid to support a form-like appearance
  - Formula-driven totals with SUMPRODUCT, IF, and ROUND
  - Print area + page setup for clean PDF export
  - Row/column grouping to hide structural scaffolding

Usage:
    from invoice import Invoice

    inv = Invoice()
    inv.build(
        company_name="Acme Corporation",
        company_details={
            "Address":  "123 Main Street, Vienna, AT 1010",
            "Phone":    "+43 1 234 5678",
            "Email":    "billing@acme.com",
            "Website":  "www.acme.com",
            "VAT ID":   "ATU12345678",
        },
        line_item_rows=12,
        tax_rate=0.20,
        currency_symbol="$",
    )
    inv.save("invoice_template.xlsx")
"""

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from xlsx_builder import XLSXBuilder, Palette, Fmt


# ── Border helpers ────────────────────────────────────────────────────────────

def _side(style="thin", color="CCCCCC"):
    return Side(style=style, color=color)

def _box(color="CCCCCC", style="thin"):
    s = _side(style, color)
    return Border(top=s, bottom=s, left=s, right=s)

def _bottom(color="CCCCCC", style="thin"):
    return Border(bottom=_side(style, color))

def _top(color="1F3864", style="medium"):
    return Border(top=_side(style, color))

THICK_BOTTOM = Border(bottom=Side(style="medium", color=Palette.DARK_NAVY))
THIN_ALL     = _box()
THIN_BOTTOM  = _bottom()
ACCENT_TOP   = _top()


class Invoice(XLSXBuilder):
    """
    Fixed-layout invoice form.

    Column grid (A–K, with a right margin at L):
      A  : left margin (narrow)
      B  : label / left content column
      C–F: Bill To / Ship To / meta content
      G  : spacer
      H–J: line item qty / price / amount
      K  : right content / amount labels
      L  : right margin (narrow)

    All user-facing input cells follow the standard blue/yellow convention.
    Formula cells are black on white.
    """

    # Column letter assignments
    _C = {
        "margin_l": "A",
        "label":    "B",
        "content":  "C",     # general left content, merges to F
        "spacer":   "G",
        "item_desc":"B",     # line item: description spans B–F
        "item_qty": "G",     # line item: quantity
        "item_rate":"H",     # line item: unit price
        "item_amt": "J",     # line item: amount (formula), merges J–K
        "total_lbl":"H",     # totals block: label
        "total_val":"J",     # totals block: value
        "margin_r": "L",
    }

    def build(self,
              company_name: str = "Your Company Name",
              company_details: dict = None,
              line_item_rows: int = 12,
              tax_rate: float = 0.10,
              currency_symbol: str = "$"):
        """
        Build the invoice workbook.

        Args:
            company_name    : Displayed prominently in the header
            company_details : Dict of label→value (Address, Phone, Email, etc.)
            line_item_rows  : Number of line item rows
            tax_rate        : Default tax rate (e.g. 0.20 for 20%)
            currency_symbol : Currency prefix shown in headers

        Returns self for chaining.
        """
        if company_details is None:
            company_details = {
                "Address": "",
                "Phone":   "",
                "Email":   "",
                "Website": "",
            }

        self._name     = company_name
        self._details  = company_details
        self._li_rows  = line_item_rows
        self._tax_rate = tax_rate
        self._currency = currency_symbol

        ws = self.add_sheet("Invoice")
        self._ws = ws
        self._setup_grid()
        self._build_header()
        self._build_address_block()
        self._build_meta_block()
        self._build_line_items()
        self._build_totals()
        self._build_footer()
        self._setup_print()

        return self

    # ── Grid setup ────────────────────────────────────────────────────────

    def _setup_grid(self):
        ws = self._ws
        ws.sheet_view.showGridLines = False

        # Column widths tuned for a classic invoice layout
        widths = {
            "A": 2,    # left margin
            "B": 14,   # labels / description start
            "C": 12,
            "D": 12,
            "E": 12,
            "F": 10,
            "G": 10,   # qty
            "H": 14,   # unit price
            "I": 2,    # internal spacer
            "J": 14,   # amount
            "K": 4,    # right pad
            "L": 2,    # right margin
        }
        self.set_col_widths(ws, widths)

    # ── Header ────────────────────────────────────────────────────────────

    def _build_header(self):
        ws = self._ws

        # Logo placeholder (rows 2–4, cols B–D)
        ws.merge_cells("B2:D4")
        c = ws["B2"]
        c.value     = "[ LOGO ]"
        c.font      = Font(name="Arial", size=11, color="BBBBBB", italic=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill      = PatternFill("solid", start_color="F5F5F5")
        for r in range(2, 5):
            ws.row_dimensions[r].height = 18

        # Company name (rows 2–3, cols E–K)
        ws.merge_cells("E2:K3")
        c = ws["E2"]
        c.value     = self._name
        c.font      = Font(name="Arial", bold=True, size=20, color=Palette.DARK_NAVY)
        c.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[2].height = 22
        ws.row_dimensions[3].height = 22

        # Company details (row 4, cols E–K)
        detail_line = "  |  ".join(
            f"{v}" for v in self._details.values() if v
        ) or "Address  |  Phone  |  Email"
        ws.merge_cells("E4:K4")
        c = ws["E4"]
        c.value     = detail_line
        c.font      = Font(name="Arial", size=9, color="666666")
        c.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[4].height = 16

        # "INVOICE" title bar (row 5)
        ws.merge_cells("B5:K5")
        c = ws["B5"]
        c.value     = "INVOICE"
        c.font      = Font(name="Arial", bold=True, size=14, color=Palette.WHITE)
        c.fill      = PatternFill("solid", start_color=Palette.DARK_NAVY)
        c.alignment = Alignment(horizontal="right", vertical="center",
                                indent=1)
        ws.row_dimensions[5].height = 24

        # Divider
        ws.row_dimensions[6].height = 6

    # ── Bill To / Ship To ─────────────────────────────────────────────────

    def _build_address_block(self):
        ws = self._ws

        for col, label in [("B", "BILL TO"), ("G", "SHIP TO")]:
            c = ws[f"{col}7"]
            c.value     = label
            c.font      = Font(name="Arial", bold=True, size=8,
                               color=Palette.WHITE)
            c.fill      = PatternFill("solid", start_color=Palette.MID_BLUE)
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        ws.merge_cells("B7:E7")
        ws.merge_cells("G7:K7")
        ws.row_dimensions[7].height = 16

        # Address input rows
        addr_fields = ["Company / Name", "Address Line 1",
                       "Address Line 2", "City, State, ZIP", "Country"]
        for i, field in enumerate(addr_fields, start=8):
            ws.row_dimensions[i].height = 18

            # Bill To
            ws.merge_cells(f"B{i}:E{i}")
            self.s.input(ws[f"B{i}"])
            ws[f"B{i}"].alignment = Alignment(horizontal="left", vertical="center",
                                              indent=1, wrap_text=False)

            # Ship To
            ws.merge_cells(f"G{i}:K{i}")
            self.s.input(ws[f"G{i}"])
            ws[f"G{i}"].alignment = Alignment(horizontal="left", vertical="center",
                                              indent=1, wrap_text=False)

        # Spacer
        ws.row_dimensions[13].height = 8

    # ── Invoice meta (Invoice #, Date, Due, PO) ───────────────────────────

    def _build_meta_block(self):
        ws = self._ws

        # Left side: "F" column label, "G–K" value
        left_meta = [
            ("Invoice #",  ""),
            ("Invoice Date", ""),
            ("Due Date",   ""),
            ("PO Number",  ""),
        ]

        meta_start = 14
        for i, (label, default) in enumerate(left_meta):
            r = meta_start + i
            ws.row_dimensions[r].height = 18

            # Label
            c = ws[f"H{r}"]
            c.value     = label
            c.font      = Font(name="Arial", bold=True, size=9,
                               color=Palette.DARK_TEXT)
            c.alignment = Alignment(horizontal="right", vertical="center")

            # Value
            ws.merge_cells(f"J{r}:K{r}")
            self.s.input(ws[f"J{r}"], value=default)
            ws[f"J{r}"].alignment = Alignment(horizontal="left", vertical="center",
                                              indent=1)

        # Spacer
        ws.row_dimensions[meta_start + len(left_meta)].height = 8

        self._line_items_start = meta_start + len(left_meta) + 1

    # ── Line items ────────────────────────────────────────────────────────

    def _build_line_items(self):
        ws  = self._ws
        row = self._line_items_start

        # Column headers
        headers = [
            ("B", "F", "DESCRIPTION"),
            ("G", "G", f"QTY"),
            ("H", "I", f"UNIT PRICE ({self._currency})"),
            ("J", "K", f"AMOUNT ({self._currency})"),
        ]
        for start, end, label in headers:
            if start != end:
                ws.merge_cells(f"{start}{row}:{end}{row}")
            c = ws[f"{start}{row}"]
            c.value     = label
            c.font      = Font(name="Arial", bold=True, size=9, color=Palette.WHITE)
            c.fill      = PatternFill("solid", start_color=Palette.DARK_NAVY)
            c.alignment = Alignment(horizontal="center" if start != "B" else "left",
                                    vertical="center", indent=1 if start == "B" else 0)
        ws.row_dimensions[row].height = 18
        row += 1

        # Line item rows
        self._li_first = row
        for i in range(self._li_rows):
            r   = row + i
            alt = i % 2 == 1
            bg  = Palette.VERY_LIGHT_BLUE if alt else Palette.WHITE

            ws.row_dimensions[r].height = 18

            # Description (B–F)
            ws.merge_cells(f"B{r}:F{r}")
            self.s.input(ws[f"B{r}"])
            if alt:
                ws[f"B{r}"].fill = PatternFill("solid", start_color="FFF9E6")

            # Qty (G)
            self.s.input(ws[f"G{r}"], fmt=Fmt.INTEGER)
            if alt:
                ws[f"G{r}"].fill = PatternFill("solid", start_color="FFF9E6")

            # Unit Price (H–I)
            ws.merge_cells(f"H{r}:I{r}")
            self.s.input(ws[f"H{r}"], fmt=Fmt.CURRENCY)
            if alt:
                ws[f"H{r}"].fill = PatternFill("solid", start_color="FFF9E6")

            # Amount = Qty * Unit Price (J–K, formula)
            ws.merge_cells(f"J{r}:K{r}")
            self.s.formula(
                ws[f"J{r}"],
                f'=IF(OR(G{r}="",H{r}=""),"",ROUND(G{r}*H{r},2))',
                fmt=Fmt.CURRENCY,
            )
            if alt:
                ws[f"J{r}"].fill = PatternFill("solid", start_color=Palette.VERY_LIGHT_BLUE)

        self._li_last = row + self._li_rows - 1
        row = self._li_last + 1

        # Thin rule below line items
        for col in ["B","C","D","E","F","G","H","I","J","K"]:
            ws[f"{col}{row}"].border = THICK_BOTTOM
        ws.row_dimensions[row].height = 4
        self._totals_start = row + 1

    # ── Totals block ──────────────────────────────────────────────────────

    def _build_totals(self):
        ws  = self._ws
        row = self._totals_start

        subtotal_formula = (
            f"=SUMPRODUCT((G{self._li_first}:G{self._li_last}<>\"\")"
            f"*(H{self._li_first}:H{self._li_last}<>\"\")"
            f"*G{self._li_first}:G{self._li_last}"
            f"*H{self._li_first}:H{self._li_last})"
        )

        def _totals_label(r, text):
            ws.merge_cells(f"H{r}:I{r}")
            c = ws[f"H{r}"]
            c.value     = text
            c.font      = Font(name="Arial", bold=True, size=10,
                               color=Palette.DARK_TEXT)
            c.alignment = Alignment(horizontal="right", vertical="center")
            ws.row_dimensions[r].height = 20

        def _totals_value(r, formula_or_value, fmt=Fmt.CURRENCY,
                          is_input=False, bold=False, highlight=False):
            ws.merge_cells(f"J{r}:K{r}")
            cell = ws[f"J{r}"]
            if is_input:
                self.s.input(cell, value=formula_or_value, fmt=fmt)
            else:
                self.s.formula(cell, formula_or_value, bold=bold, fmt=fmt)
            if highlight:
                cell.fill = PatternFill("solid", start_color=Palette.TOTAL_FILL)

        # Subtotal
        _totals_label(row, "Subtotal")
        _totals_value(row, subtotal_formula)
        self._subtotal_row = row
        row += 1

        # Discount
        _totals_label(row, "Discount")
        _totals_value(row, 0, fmt=Fmt.CURRENCY, is_input=True)
        self._discount_row = row
        row += 1

        # Tax rate (input)
        _totals_label(row, f"Tax Rate")
        _totals_value(row, self._tax_rate, fmt=Fmt.PCT, is_input=True)
        self._tax_rate_row = row
        row += 1

        # Tax amount (formula)
        _totals_label(row, "Tax Amount")
        _totals_value(
            row,
            f"=ROUND((J{self._subtotal_row}-J{self._discount_row})*J{self._tax_rate_row},2)",
        )
        self._tax_amt_row = row
        row += 1

        # Total Due
        ws.row_dimensions[row].height = 24
        ws.merge_cells(f"H{row}:I{row}")
        c = ws[f"H{row}"]
        c.value     = "TOTAL DUE"
        c.font      = Font(name="Arial", bold=True, size=12, color=Palette.WHITE)
        c.fill      = PatternFill("solid", start_color=Palette.DARK_NAVY)
        c.alignment = Alignment(horizontal="right", vertical="center")

        ws.merge_cells(f"J{row}:K{row}")
        c = ws[f"J{row}"]
        c.value  = (
            f"=ROUND(J{self._subtotal_row}"
            f"-J{self._discount_row}"
            f"+J{self._tax_amt_row},2)"
        )
        c.font      = Font(name="Arial", bold=True, size=12, color=Palette.WHITE)
        c.fill      = PatternFill("solid", start_color=Palette.DARK_NAVY)
        c.number_format = Fmt.CURRENCY
        c.alignment = Alignment(horizontal="right", vertical="center")
        self._total_row = row
        row += 1
        self._footer_start = row + 1

    # ── Footer ────────────────────────────────────────────────────────────

    def _build_footer(self):
        ws  = self._ws
        row = self._footer_start

        # Divider
        ws.merge_cells(f"B{row}:K{row}")
        ws[f"B{row}"].border = THICK_BOTTOM
        ws.row_dimensions[row].height = 4
        row += 1

        # Payment Terms + Bank Details side by side
        footer_sections = [
            ("B", "E", "PAYMENT TERMS"),
            ("G", "K", "BANK / PAYMENT DETAILS"),
        ]
        for start, end, label in footer_sections:
            ws.merge_cells(f"{start}{row}:{end}{row}")
            c = ws[f"{start}{row}"]
            c.value     = label
            c.font      = Font(name="Arial", bold=True, size=8, color=Palette.WHITE)
            c.fill      = PatternFill("solid", start_color=Palette.MID_BLUE)
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 16
        row += 1

        # Input rows for each section
        for _ in range(4):
            ws.merge_cells(f"B{row}:E{row}")
            self.s.input(ws[f"B{row}"])
            ws.merge_cells(f"G{row}:K{row}")
            self.s.input(ws[f"G{row}"])
            ws.row_dimensions[row].height = 18
            row += 1

        # Notes
        row += 1
        ws.merge_cells(f"B{row}:K{row}")
        subhdr = ws[f"B{row}"]
        subhdr.value     = "NOTES"
        subhdr.font      = Font(name="Arial", bold=True, size=8, color=Palette.WHITE)
        subhdr.fill      = PatternFill("solid", start_color=Palette.MID_BLUE)
        subhdr.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 16
        row += 1

        for _ in range(3):
            ws.merge_cells(f"B{row}:K{row}")
            self.s.textarea(ws[f"B{row}"])
            ws[f"B{row}"].alignment = Alignment(horizontal="left", vertical="center",
                                                indent=1, wrap_text=True)
            ws.row_dimensions[row].height = 20
            row += 1

        # Thank-you line
        row += 1
        ws.merge_cells(f"B{row}:K{row}")
        c = ws[f"B{row}"]
        c.value     = "Thank you for your business."
        c.font      = Font(name="Arial", italic=True, size=9, color="888888")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 20
        self._last_row = row

    # ── Print / page setup ────────────────────────────────────────────────

    def _setup_print(self):
        ws = self._ws

        # Print area: B2 to K(last_row)
        ws.print_area = f"A2:L{self._last_row}"

        # Page setup for A4
        ws.page_setup.paperSize  = ws.PAPERSIZE_A4
        ws.page_setup.orientation = "portrait"
        ws.page_setup.fitToPage  = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

        # Margins (in inches)
        ws.page_margins.left   = 0.4
        ws.page_margins.right  = 0.4
        ws.page_margins.top    = 0.5
        ws.page_margins.bottom = 0.5
